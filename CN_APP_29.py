import glob
import os,shutil
from shutil import make_archive
import tkinter as tk
from tkinter import ttk,messagebox
import base64
from tkinter import *
import tkinter.font as font
import pyautogui as pg

import time
from datetime import date
from datetime import datetime,timedelta

from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as xlimg
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

import win32api,win32con
from threading import Thread
import dicom as dcm
import cv2
import numpy as np
import math
import sys
import codecs
import pywinauto
from pywinauto import application

import sys, os
import sqlite3 as sq

import qrcode
from PIL import Image
import json

import socket
import pickle

import matplotlib.pyplot as plt







def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


i=int()
i=0
j=int()
j=1
k=int()
k=1
j_k = int()
j_k = 1

timing = datetime.now()
if 21<=timing.hour<24 :
    time_value=datetime.now()+timedelta(1)
    print('this is time value = ',time_value)
else:
    time_value=datetime.now()


time_value=time_value.strftime("%d-%m-%y")
print(time_value)

max_line = 22# MAX8LINES +1
result_list=[]
interpreted = False
quit_var = False
ARCHIVE=int()
ARCHIVE =0



NO_border = Border(left=Side(border_style=None), 
                         right=Side(border_style=None), 
                         top=Side(border_style=None), 
                         bottom=Side(border_style=None))

""""this code should be working on top of YXLON image 3500 software,  it generates an excel file and write the repport while doing the interpretation of dcm films"""
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.getcwd()
    return os.path.join(base_path, relative_path)


##def resource_path(relative_path):
##    """ Get absolute path to resource, works for dev and for PyInstaller """
##    if getattr(sys, 'frozen', False):
##        base_path = sys._MEIPASS
##    else:
##        base_path = os.getcwd()
##    return os.path.join(base_path, relative_path)

####################################################################################################################################################################################################################################################################################################################################################################################################################
####################################################################################################################################################################################################################################################################################################################################################################################################################
####################################################################################################################################################################################################################################################################################################################################################################################################################


class load_path:
    def load_element(self, element):
        self.jsn_list="projet.json"
        try:
          openfile = open(self.jsn_list, 'r')
          jsn_conten = json.load(openfile)
        except:
            outfile = open(self.jsn_list, "w")
            if self.jsn_list=="projet.json":
                J_dict={"projet":"Rgz-2",
                        "Nuance":"x70  MPSL 2",
                        "Diameter":"1016 mm",
                        "Epaisseur":"12,70 mm",
                        "ip_address":"192.168.2.4",
                        "values":[-60,-30,-20,-15,-10,0,10,15,20,30,60,80],
                        "values_wait":25,
                       
                        
                        "defaut_AREP":["AA","BA","BU","DL","F","SCVE"],
                        "defaut_AMEULER":["AA","BA","BU","DL","F","SCVE"],
                        "DEFAULT_ACH":['AA/D','BA/D','BU/D','DL/D','CF/D','MM/D','AA/F','BA/F','BU/F','DL/F','CF/F','MM/F'],
                        "defaut_OK":["AA","BA","BU","DL","F","SCVE"],
                        
                        
                        "Classes":["AA_BA","DL","BU","AN,MMC","F","WELD_IMPERFECTION"],
                        "LISTE_DES_POST":[5,13,13,21,21,5],
                        
                        "LISTE_operateurs":['BOUZID YASSINE','BOUZID YASSINE','BOUZID YASSINE'],
                        #"LISTE_Projets":['CEEG KD/AL','Rgz-2'],
                        
                        "PATH_PROJET": r"C:\Users\111\Desktop\1750",
                        "PATH_CLIENT": r"C:\Users\111\Desktop\client",
                        "PATH_SCOPIE": r"//YASSINE-PC/shared folder/SCOPY",
                        "path_fab"   : r"C:\Users\YASSINE\Desktop\SCRIPTS\fabrication",
                        "PATH_prog"  : r"C:\Program Files\YXLON\Y.Image 3500\IMAGE_3500.exe",
                        "bits":8
                        }   
            json_object = json.dumps(J_dict, indent=2)
            outfile.write(json_object)
            jsn_conten = J_dict
        return jsn_conten[element]
    
P = load_path()                  

thread_condition=True
ask_condition=True
Start=True
threa_3_is_not_started = True

def check_live(event):
    global show_condition,Start,ask_condition,threa_3_is_not_started
    print("live stream is clicked",var12.get())

    ask_condition =True
    if var12.get():
        show_condition=False
    else:
        show_condition=True
       
        
    if threa_3_is_not_started:
        t3.start()
        threa_3_is_not_started = False

        
def RUN_stream():
    global show_condition,ask_condition,Start
    print("threa run stream is started")
    while thread_condition:
            try:
                c = socket.socket()
                c.connect((P.load_element("ip_address"), 9999))
                msg = c.recv(1000160)
                gray_image = pickle.loads(msg)
                gray_image = cv2.cvtColor(gray_image, cv2.COLOR_BGRA2BGR)
                if show_condition:
                    cv2.imshow('Live Streaming', gray_image)
                    if cv2.waitKey(25) & 0xFF == ord('q'):
                        cv2.destroyWindow('Live Streaming')
                else:
                    cv2.destroyWindow('Live Streaming')
                    
            except Exception as e:
                if ask_condition :
                    confirmation = messagebox.askquestion("Connection faild!","RX1 STREAM IS NOT CONNECTED!!\n Would you like to try again?")
                    cv2.destroyWindow('Live Streaming')
                    if confirmation== "yes":
                        pass
                    else:
                        ask_condition =False
                        LIVE_SCOPY.deselect()
                        show_condition=False
                        print("live stream checkbox --",var12.get())
                       
                        
          
def quitt():
    global thread_condition
    thread_condition = False
    sys.exit()
    root.quit()

        
####################################################################################################################################################################################################################################################################################################################################################################################################################
####################################################################################################################################################################################################################################################################################################################################################################################################################



def covert_to_excel(listing):
    listing =  (str(listing)).replace('[','')
    listing =  listing.replace(']','')
    listing =  listing.replace("'","")
    listing =  listing.replace("\n","")
    return listing

def covert_PATH(listing):
    listing =  (str(listing)).replace('/','_')
    listing =  listing.replace("\\", '_')
    listing =  listing.replace("<","")
    listing =  listing.replace(">","")                           
    listing =  listing.replace("?","")
    listing =  listing.replace(":","")
    listing =  listing.replace("*","")
    listing =  listing.replace('"',"")
    listing =  listing.replace("|","")
    return listing

def creating_the_zip_file():
    dir_name="zip folder"
    output_filename="Images"
    if not os.path.exists(dir_name):
            os.makedirs(dir_name) 
    list_of_filmes= glob.glob(newpath+'/*.jpg')
    for each_film in list_of_filmes:
        shutil.copy(each_film, dir_name)
    shutil.make_archive(output_filename, 'zip', dir_name)
    
    for each_jpg in glob.glob(dir_name+'/*.jpg'):
        os.remove(each_jpg)

def convertToBinaryData(filename):
    # Convert digital data to binary format
    with open(filename, 'rb') as file:
        blobData = file.read()
    return blobData


def rq_genrator(string,image_name_and_link):
    img = qrcode.make(string)
    img.save(f"{image_name_and_link}.jpg")
    img = Image.open(f"{image_name_and_link}.jpg")
    img = img.convert("RGBA")

    pixdata = img.load()

    width, height = img.size
    for y in range(height):
        for x in range(width):
            if pixdata[x, y] == (255, 255, 255, 255):
                pixdata[x, y] = (255, 255, 255, 0)
    new_width  = 150
    new_height = 180
    img = img.resize((new_width, new_height), Image.ANTIALIAS)
    img.save("img2.jpg", "PNG")
    
    wb = load_workbook('QR_template.xlsx')
    wb.save("QR_print.xlsx")
        
    wb = load_workbook("QR_print.xlsx")   
    ws = wb.active
    try:
        rq_image =xlimg("img2.jpg")  
        ws.add_image(rq_image,'H8')
        wb.save("QR_print.xlsx")
        
    except Exception as e:
        print("eeeeeeeee",e)
    win32api.ShellExecute(
        0,
        "print",
        ("QR_print.xlsx"),
        None,
        ".",
        0
        )
    return img
    

####################################################################################################################################################################################################################################################################################################################################################################################################################


def adding_numbers_toslider(event):
    if event.delta > 0:
       scan_slider_value =int(scan_slider.get())+5
    else:
       scan_slider_value =int(scan_slider.get())-5

    scan_slider.delete(0,"end")
    scan_slider.insert(0,scan_slider_value)
    slider_update(event)

def slider_update(event):        
    name_file =list_film.get(ANCHOR)
    name_film = name_file[6:]
    save_path = f"{newpath}\{name_file}"
    img = dcm.read_file(save_path)
    cv2.imwrite("img1.jpg", (np.median(img.pixel_array)+int(scan_slider.get())-img.pixel_array))
    image_read=cv2.imread("img1.jpg")
    cv2.imshow("IMAGE",image_read)
    
    
def read_dicom(event):
    name_file =list_film.get(ANCHOR)
    name_film = name_file[6:]
    save_path = f"{newpath}\{name_file}"
    imgdcm = dcm.read_file(save_path)
    net = cv2.dnn.readNet("yolov3_training_last-5.weights", "yolov3_testing7classes.cfg")
    valuess=P.load_element("values")#load_list("values.txt")#[-60,-30,-20,-15,-10,0,10,15,20,30,60]
    values_wait = P.load_element("values_wait")#load_list("values_wait.txt")
    if "IQI" in name_file:
        cv2.imwrite("img1.jpg", (np.median(imgdcm.pixel_array)+ int(scan_slider.get())-imgdcm.pixel_array))
        img = cv2.imread("img1.jpg")
        img_gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        template = cv2.imread("iqi_template1.jpg")
        template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)

        # Perform template matching using cv2.matchTemplate()
        result = cv2.matchTemplate(img_gray, template_gray, cv2.TM_CCOEFF_NORMED)
        # Find the location of the maximum correlation value
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

        # Draw a rectangle around the matched template
        top_left = max_loc
        print("max_val max_loc",max_val, max_loc)
        bottom_right = (top_left[0] + template.shape[1], top_left[1] + template.shape[0])
        
        if max_val>=.5:
            # Specify the rectangle properties
            length =template.shape[1]*3          # 40 mm 
            width = template.shape[0]*1.5        # 30 mm 
            angle= 35
            size = (length,width)
           
            # Calculate the cosine and sine of the angle
            cos_value = math.cos(math.radians(angle))
            sin_value = math.sin(math.radians(angle))
            
            top_left_ = (top_left[0]+template.shape[0]/2,top_left[1]-template.shape[1]/2)
            # Calculate the top-left corner
            top_left_rectangle_x = top_left_[0] + size[0]/2*cos_value - size[1]/2*sin_value
            top_left_rectangle_y = top_left_[1] + size[0]/2*sin_value + size[1]/2*cos_value
            
            top_left_rectangle = (top_left_rectangle_x,top_left_rectangle_y) 
            print("top_left_rectangle",top_left_rectangle)
            

            rectangle = top_left_rectangle ,size,angle #(100, 100), (200, 200), 45

            # Get the four corner points of the rectangle
            box = cv2.boxPoints(rectangle)
            print("the boxes1 ",box)

            # Convert the points to integers
            box = np.int0(box)
            
            # Draw the polygon
            cv2.polylines(img, [box], True, (255, 0, 0), 1)
            
            # Get the bounding rectangle of the polygon region
            print("the boxes ",box)
            
            x, y, w, h = cv2.boundingRect(box)
            print("x, y, w, h",x, y, w, h)
            
            # Get the rotation matrix
            M = cv2.getRotationMatrix2D(top_left, angle, scale = 1)
            print("M",M)
            # Perform the affine transformation
            rotated_image = cv2.warpAffine(img, M, (1100, 1100))

            
            cropped_image = rotated_image[(y-40):(y-(h+450)),(x+83):(x+23+w)]
            # Apply the filter using cv2.filter2D
             # Define the filter kernel
            kernel = np.array([[14,   14,  14],
                               [.2,  3, .2],
                               [-14, -14, -14]])
            cropped_image = cv2.filter2D(cropped_image, -1, kernel)

            gray = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
           

            
            #kernel = np.ones((3,3),np.float32)/25
            #dst = cv2.filter2D(gray,-1,kernel)

            #Apply Gaussian smoothing
            kernel_size = (5, 5)
            dst = gray

            
            print('the shape of  dst',dst.shape)
            row = gray[gray.shape[0]-10,:]
            max_value = max(row)
            max_index = [i for i, value in enumerate(row) if value == max_value][0]
            
            rect = dst[:,max_index-50:max_index+50]
            cv2.imshow("ROI",rect)
            canny_edges = cv2.Canny(rect, 120, 150)
            cv2.imshow("canny_edges",canny_edges)
                        
            
            print("te maximum value and index ",(max_value,max_index))
            
            #Define the start and end points of the line
            start_point1 = (max_index, 0)
            end_point1 = (max_index, h)
            
            start_point2 = (max_index+30, 0)
            end_point2 = (max_index+30, h)
            
            start_point3 = (max_index-30, 0)
            end_point3 = (max_index-30, h)

            # Define the color and thickness of the line
            
            color1 = (255, 0, 0) #  blue
            color2 = (0, 0, 255) #  red
            color3 = (0, 255, 0) #  green
            thickness = 1
            # Draw the line on the image
            cv2.line(cropped_image, start_point1, end_point1, color1, thickness)
            cv2.line(cropped_image, start_point2, end_point2, color2, thickness)
            cv2.line(cropped_image, start_point3, end_point3, color3, thickness)
            
            
            column1 = dst[10:-10,max_index]
            column2 = dst[10:-10,max_index-30]
            column3 = dst[10:-10,max_index+30]
            
            indexes = [i for i, value in enumerate(column1)]
            import scipy
            from scipy.signal import savgol_filter

            column1 = np.array(column1)
            column1 = savgol_filter(column1, 9, 1)
            
            column2 = np.array(column2)
            column2 = savgol_filter(column2, 9, 1)
            
            column3 = np.array(column3)
            column3 = savgol_filter(column3, 9, 1)
            
            max_column= max(-column1)
            max_index = [i for i, value in enumerate(-column1) if value == max_column][0]
            print("index and max column --------", max_index ,max_column )
            

            #plt.plot(-column1,-column2,-column3)
            plt.plot(indexes, column1, 'b') # plotting t, a separately
            plt.plot(indexes, column2, 'r') # plotting t, a separately
            plt.plot(indexes, column3, 'g') # plotting t, a separately 
            # Add labels and title to the plot
            plt.xlabel('Index')
            plt.ylabel('Value')
            plt.title('Column Plot')
            #plt.savefig('my_plot.png')
            plt.savefig('iqi.png')

            # Show the plot
            plt.show()
            #cropped_image = cv2.Canny(cropped_image, 100, 200)
           


            
            # Save the cropped image
            cv2.imwrite('cropped2.jpg', cropped_image)
            cropped_image = cv2.transpose(cropped_image)

            # Flip the image horizontally
            cropped_image = cv2.flip(cropped_image, 1)
            cv2.imshow("cropped_image", cropped_image)
            cv2.waitKey(2000)

                        
            cv2.rectangle(img, top_left, bottom_right, (0, 0, 255), 2)
            
        else:
           cv2.putText(img, "DO NOT MATCH", (100, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255,0), 2)

#

        # Show the result
        #cv2.imshow("Result", img)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows()
        

        return "break"

    
    for valuei in valuess:
        
        cv2.imwrite("img1.jpg", (np.median(imgdcm.pixel_array)+int(valuei)-imgdcm.pixel_array))
        # Load Yolo
        #net = cv2.dnn.readNet("yolov3_training_last-5.weights", "yolov3_testing7classes.cfg")
        layer_names = net.getLayerNames()
        output_layers = [layer_names[i- 1] for i in net.getUnconnectedOutLayers()]

        # Loading image
        img = cv2.imread("img1.jpg")
        #img = cv2.imread(r"\\YASSINE-PC\test\A7540-K1_BUC A REP.jpg")
        height, width, channels = img.shape

        # Detecting objects
        blob = cv2.dnn.blobFromImage(img, 0.00392, (416, 416), (0, 0, 0), True, crop=True)
        net.setInput(blob)
        outs = net.forward(output_layers)
        
        classes = P.load_element("Classes")#load_list("Classes.txt")
        
        # Showing informations on the screen
        class_ids = []
        confidences = []
        boxes = []
        for out in outs:
            for detection in out:
                scores = detection[5:]
                class_id = np.argmax(scores)
                confidence = scores[class_id]
                if confidence > 0:
                    # Object detected
                    print("THE CLASS ID IS =====-------======",class_id)
                    center_x = int(detection[0] * width)
                    center_y = int(detection[1] * height)
                    w = int(detection[2] * width)
                    h = int(detection[3] * height)

                    # Rectangle coordinates
                    x = int(center_x - w / 2)
                    y = int(center_y - h / 2)

                    boxes.append([x, y, w, h])
                    confidences.append(float("{:.2f}".format(confidence)))
                    class_ids.append(class_id)
                    #font = cv2.FONT_HERSHEY_COMPLEX_SMALL
                    font = cv2.FONT_HERSHEY_PLAIN
        lst = []
        indexes = cv2.dnn.NMSBoxes(boxes, confidences,0.1,.3)
        l=0
        #print('indexes= ',indexes)
        #print('confidences= ',confidences)
        #print('boxes= ',boxes)
        for i in range(len(boxes)):
                if i not in indexes:
                    l+=1
                if i in indexes:
                    x, y, w, h = boxes[i]
                    defct_dimention = (math.sqrt(w ** 2 + h ** 2)) * .1#the repport between the distance (object---detector) over (object---source)
                    print("defect dimention=" + str(defct_dimention))
                    lst.append("{:.2f}".format(defct_dimention))
                    label = f"{classes[class_ids[i]]}-" + lst[i-l] + "mm" #+str(float("{:.2f}".format(confidences[l]*100)))+"%"
                    coleur = (255,0,0)
                    if float(lst[i-l])>3:
                        coleur = (0,0,255)
                        
                    cv2.rectangle(img, (x, y), (x + w, y + h),coleur , 2)
                    cv2.putText(img, label, (x-100, y-20 ), font, 1, (0, 255, 0), 2)
        
        devlabel.configure( text="defect dimention=\n"+ str(lst).replace(',','\n')+ " mm\n"+ str(confidences).replace(',','\n') + " %",bg="yellow")
        cv2.imwrite(f"imgcv{valuei}.jpg", img)
        cv2.imshow("SCANNED IMAGE",img)
        if cv2.waitKey(values_wait) & 0xFF == ord('q'):
            cv2.destroyAllWindows()

##    print("Done!!")
##
##
##def read_dicom(event):
##    name_file =list_film.get(ANCHOR)
##    
##    
##    name_film = name_file[6:]
##    save_path = f"{newpath}\{name_file}"
##    imgdcm = dcm.read_file(save_path)
##    net = cv2.dnn.readNet("yolov3_training_last-5.weights", "yolov3_testing7classes.cfg")
##    valuess=P.load_element("values")#load_list("values.txt")#[-60,-30,-20,-15,-10,0,10,15,20,30,60]
##    values_wait = P.load_element("values_wait")#load_list("values_wait.txt")
####    
##    if "IQI" in name_file:
##        cv2.imwrite("img1.jpg", (np.median(imgdcm.pixel_array)+ int(scan_slider.get())-imgdcm.pixel_array))
##        img = cv2.imread("img1.jpg")
##        img_gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
##        template = cv2.imread("iqi_template1.jpg")
##        template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
##
##        # Perform template matching using cv2.matchTemplate()
##        result = cv2.matchTemplate(img_gray, template_gray, cv2.TM_CCOEFF_NORMED)
##        # Find the location of the maximum correlation value
##        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
##
##        # Draw a rectangle around the matched template
##        top_left = max_loc
##        print("max_val max_loc",max_val, max_loc)
##        bottom_right = (top_left[0] + template.shape[1], top_left[1] + template.shape[0])
##        
##        if max_val>=.5:
##            # Specify the rectangle properties
##            length =template.shape[1]*3          # 40 mm 
##            width = template.shape[0]*1.5        # 30 mm 
##            angle= 35
##            size = (length,width)
##           
##            # Calculate the cosine and sine of the angle
##            cos_value = math.cos(math.radians(angle))
##            sin_value = math.sin(math.radians(angle))
##            top_left_ = (top_left[0]+template.shape[0]/2,top_left[1]-template.shape[1]/2)
##            # Calculate the top-left corner
##            top_left_rectangle_x = top_left_[0] + size[0]/2*cos_value - size[1]/2*sin_value
##            top_left_rectangle_y = top_left_[1] + size[0]/2*sin_value + size[1]/2*cos_value
##            top_left_rectangle = (top_left_rectangle_x,top_left_rectangle_y) 
##            print("top_left_rectangle",top_left_rectangle)
##            
##
##            rectangle = top_left_rectangle ,size,angle #(100, 100), (200, 200), 45
##
##            # Get the four corner points of the rectangle
##            box = cv2.boxPoints(rectangle)
##            print("the boxes1 ",box)
##
##            # Convert the points to integers
##            box = np.int0(box)
##            
##            # Draw the polygon
##            cv2.polylines(img, [box], True, (255, 0, 0), 1)
##            
##            # Get the bounding rectangle of the polygon region
##            print("the boxes ",box)
##            
##            x, y, w, h = cv2.boundingRect(box)
##            print("x, y, w, h",x, y, w, h)
##            
##            # Get the rotation matrix
##            M = cv2.getRotationMatrix2D(top_left, angle, scale = 1)
##            print("M",M)
##            # Perform the affine transformation
##            rotated_image = cv2.warpAffine(img, M, (1100, 1100))
##
##
##            
##            
##            cropped_image = rotated_image[(y-40):(y-(h+450)),(x+83):(x+23+w)]
##            gray = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
##            # Define the filter kernel
##            kernel = np.array([[14,   14,  14],
##                               [.2,  3, .2],
##                               [-14, -14, -14]])
##
##            # Apply the filter using cv2.filter2D
##            cropped_image = cv2.filter2D(cropped_image, -1, kernel)
##
##            #kernel = np.ones((3,3),np.float32)/25
##            #dst = cv2.filter2D(gray,-1,kernel)
##
##            #Apply Gaussian smoothing
##            kernel_size = (5, 5)
##            dst = cv2.GaussianBlur(gray, kernel_size, 0)
##
##            
##            print('the shape of  dst',dst.shape)
##            row = dst[dst.shape[0]-10,:]
##            max_value = max(row)
##            max_index = [i for i, value in enumerate(row) if value == max_value][0]
##            print("te maximum value and index ",(max_value,max_index))
##            
##            #Define the start and end points of the line
##            start_point = (max_index, 0)
##            end_point = (max_index, h)
##
##            # Define the color and thickness of the line
##            color = (0, 255, 0) # Blue color
##            thickness = 2
##            # Draw the line on the image
##            cv2.line(cropped_image, start_point, end_point, color, thickness)
##            
##            
##            column = dst[5:-5,max_index]
##            import scipy
##            from scipy.signal import savgol_filter
##
##            column = np.array(column)
##            column = savgol_filter(column, 33, 1)
##            
##
##            plt.plot(-column)
##            # Add labels and title to the plot
##            plt.xlabel('Index')
##            plt.ylabel('Value')
##            plt.title('Column Plot')
##            plt.savefig('my_plot.png')
##            plt.savefig('iqi.png')
##
##            # Show the plot
##            plt.show()
##            #cropped_image = cv2.Canny(cropped_image, 100, 200)
##           
##
##
##            
##            # Save the cropped image
##            cv2.imwrite('cropped2.jpg', cropped_image)
##            cv2.imshow("cropped_image", cropped_image)
##            cv2.waitKey(2000)
##
##                        
##            cv2.rectangle(img, top_left, bottom_right, (0, 0, 255), 2)
##            
##        else:
##           cv2.putText(img, "DO NOT MATCH", (100, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255,0), 2)
##
###
##
##        # Show the result
##        cv2.imshow("Result", img)
##        cv2.waitKey(0)
##        cv2.destroyAllWindows()
##        
##
##        return "break"
##    
##    for valuei in valuess:
##        
##        cv2.imwrite("img1.jpg", (np.median(imgdcm.pixel_array)+int(valuei)-imgdcm.pixel_array))
##        # Load Yolo
##        #net = cv2.dnn.readNet("yolov3_training_last-5.weights", "yolov3_testing7classes.cfg")
##        layer_names = net.getLayerNames()
##        output_layers = [layer_names[i[0] - 1] for i in net.getUnconnectedOutLayers()]
##
##        # Loading image
##        img = cv2.imread("img1.jpg")
##        
##        #img = cv2.imread(r"\\YASSINE-PC\test\A7540-K1_BUC A REP.jpg")
##        height, width, channels = img.shape
##
##        # Detecting objects
##        blob = cv2.dnn.blobFromImage(img, 0.00392, (416, 416), (0, 0, 0), True, crop=True)
##        net.setInput(blob)
##        outs = net.forward(output_layers)
##        
##        classes = P.load_element("Classes")#load_list("Classes.txt")
##        
##        # Showing informations on the screen
##        class_ids = []
##        confidences = []
##        boxes = []
##        for out in outs:
##            for detection in out:
##                scores = detection[5:]
##                class_id = np.argmax(scores)
##                confidence = scores[class_id]
##                if confidence > 0:
##                    # Object detected
##                    print("THE CLASS ID IS =====-------======",class_id)
##                    center_x = int(detection[0] * width)
##                    center_y = int(detection[1] * height)
##                    w = int(detection[2] * width)
##                    h = int(detection[3] * height)
##
##                    # Rectangle coordinates
##                    x = int(center_x - w / 2)
##                    y = int(center_y - h / 2)
##
##                    boxes.append([x, y, w, h])
##                    confidences.append(float("{:.2f}".format(confidence)))
##                    class_ids.append(class_id)
##                    #font = cv2.FONT_HERSHEY_COMPLEX_SMALL
##                    font = cv2.FONT_HERSHEY_PLAIN
##        lst = []
##        indexes = cv2.dnn.NMSBoxes(boxes, confidences,0.1,.3)
##        l=0
##        #print('indexes= ',indexes)
##        #print('confidences= ',confidences)
##        #print('boxes= ',boxes)
##        for i in range(len(boxes)):
##                if i not in indexes:
##                    l+=1
##                if i in indexes:
##                    x, y, w, h = boxes[i]
##                    defct_dimention = (math.sqrt(w ** 2 + h ** 2)) * .1#the repport between the distance (object---detector) over (object---source)
##                    print("defect dimention=" + str(defct_dimention))
##                    lst.append("{:.2f}".format(defct_dimention))
##                    label = f"{classes[class_ids[i]]}-" + lst[i-l] + "mm" #+str(float("{:.2f}".format(confidences[l]*100)))+"%"
##                    coleur = (255,0,0)
##                    if float(lst[i-l])>3:
##                        coleur = (0,0,255)
##                        
##                    cv2.rectangle(img, (x, y), (x + w, y + h),coleur , 2)
##                    cv2.putText(img, label, (x-100, y-20 ), font, 1, (0, 255, 0), 2)
##        
##        devlabel.configure( text="defect dimention=\n"+ str(lst).replace(',','\n')+ " mm\n"+ str(confidences).replace(',','\n') + " %",bg="yellow")
##        cv2.imwrite(f"imgcv{valuei}.jpg", img)
##        cv2.imshow("SCANNED IMAGE",img)
##        if cv2.waitKey(values_wait) & 0xFF == ord('q'):
##            cv2.destroyAllWindows()
##
##    print("ffDone!!")
####################################################################################################################################################################################################################################################################################################################################################################################################################
#    


def read_dicom_reporter(newpath):
    name_file =list_film.get(ANCHOR)
    name_film = name_file[6:]
    save_path = f"{newpath}\{name_file}"
    
    net = cv2.dnn.readNet("yolov3_training_last-5.weights", "yolov3_testing7classes.cfg")
    layer_names = net.getLayerNames()
    output_layers = [layer_names[i[0] - 1] for i in net.getUnconnectedOutLayers()]
    
    valuess=P.load_element("values")#load_list("values.txt")#[-60,-30,-20,-15,-10,0,10,15,20,30,60]
    values_wait = P.load_element("values_wait")#load_list("values_wait.txt")
    for film in os.listdir(newpath) :
        
            boxessss=[]
            confidencesss=[]
            class_idsss=[]
            imgdcm = dcm.read_file(f"{newpath}\{film}")
            for valuei in valuess:
                
                cv2.imwrite("img1.jpg", (np.median(imgdcm.pixel_array)+int(valuei)-imgdcm.pixel_array))
                img = cv2.imread("img1.jpg")
                os.remove("img1.jpg")
                
                # Detecting objects
                blob = cv2.dnn.blobFromImage(img, 0.00392, (416, 416), (0, 0, 0), True, crop=True)
                net.setInput(blob)
                outs = net.forward(output_layers)
                
                classes = P.load_element("Classes")#load_list("Classes.txt")
                
                # Showing informations on the screen
                class_ids = []
                confidences = []
                boxes = []
                for out in outs:
                    for detection in out:
                        scores = detection[5:]
                        class_id = np.argmax(scores)
                        confidence = scores[class_id]
                        if confidence > 0:
                            # Object detected
                            print("THE CLASS ID IS =====-------======",class_id)
                            center_x = int(detection[0] * 1000)
                            center_y = int(detection[1] * 1000)
                            w = int(detection[2] * 1000)
                            h = int(detection[3] * 1000)

                            # Rectangle coordinates
                            x = int(center_x - w / 2)
                            y = int(center_y - h / 2)

                            boxes.append([x, y, w, h])
                            confidences.append(float("{:.2f}".format(confidence)))
                            class_ids.append(class_id)
                            #font = cv2.FONT_HERSHEY_COMPLEX_SMALL
                            font = cv2.FONT_HERSHEY_PLAIN
                lst = []
                indexes = cv2.dnn.NMSBoxes(boxes, confidences,0.1,.3)
                l=0
                #print('indexes= ',indexes)
                #print('confidences= ',confidences)
                #print('boxes= ',boxes)
##                for i in range(len(boxes)):
##                        if i not in indexes:
##                            l+=1
##                        if i in indexes:
##                            x, y, w, h = boxes[i]
##                            defct_dimention = (math.sqrt(w ** 2 + h ** 2)) * .1#the repport between the distance (object---detector) over (object---source)
##                            print("defect dimention=" + str(defct_dimention))
##                            lst.append("{:.2f}".format(defct_dimention))
##                            label = f"{classes[class_ids[i]]}-" + lst[i-l] + "mm" #+str(float("{:.2f}".format(confidences[l]*100)))+"%"
##                            coleur = (255,0,0)
##                            if float(lst[i-l])>3:
##                                coleur = (0,0,255)
##                                
##                            cv2.rectangle(img, (x, y), (x + w, y + h),coleur , 2)
##                            cv2.putText(img, label, (x-100, y-20 ), font, 1, (0, 255, 0), 2)
##                
##                devlabel.configure( text="defect dimention=\n"+ str(lst).replace(',','\n')+ " mm\n"+ str(confidences).replace(',','\n') + " %",bg="yellow")
            #cv2.imwrite(f"{newpath}/{film}-{valuei}.jpg", img)
            #cv2.imshow("SCANNED IMAGE",img)
            #os.remove(f"{newpath}/{film}-{valuei}.jpg")
            
                boxessss.append(boxes)#[box for i,box in enumerate (boxes) if box[i]-box[i+1]>0 else box=""])
                confidencesss.append(confidences)
                class_idsss.append([classes[i]for i in class_ids])
            with open(f'{newpath}/{film[:-8]}.txt', 'w') as f:
                f.write(f'boxessss = {boxessss}\nconfidencesss = {confidencesss}\nclass_idsss = {class_idsss}\n')
                #if cv2.waitKey(values_wait) & 0xFF == ord('q'):
                    #cv2.destroyAllWindows()

    print("ssDone!!")

##    
####################################################################################################################################################################################################################################################################################################################################################################################################################

##    
####################################################################################################################################################################################################################################################################################################################################################################################################################
####################################################################################################################################################################################################################################################################################################################################################################################################################
####################################################################################################################################################################################################################################################################################################################################################################################################################

def create_atable(NAME_DATABASE,NAME_TABLE,NAME_REPORT,col1,col2,col3,col4,col5,col6):
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"CREATE TABLE IF NOT EXISTS {NAME_TABLE} ({NAME_REPORT} text, {col1} text, {col2}  integer, {col3}  text, {col4} text , {col5} text, {col6} blob)")
    cnn.commit()
    cnn.close()
    print("done!")

def insert_one_record(NAME_DATABASE,NAME_TABLE,one_record):
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"INSERT INTO {NAME_TABLE} VALUES(?,?,?,?,?,?,?)",one_record)
    cnn.commit()
    cnn.close()
    print("done!")
def move_iqi_into_folder():
        list_of_filmes= glob.glob(newpath+'/*')
        print("list_of_filmes",list_of_filmes)
        interpreted_iqi_path = os.path.dirname(newpath) + "/IQI-INTERPRTER"
        
        if not os.path.exists(interpreted_iqi_path):
            os.makedirs(interpreted_iqi_path)
            
        for each_film in list_of_filmes:
            if os.path.exists(each_film):
                os.remove(dst)
            
            shutil.move(each_film, interpreted_iqi_path)
                
def tube_finished():
    global i,j,file12,k,result_list,ws,j_k,name_folder
##    path = covert_to_excel(load_list("PATH_PROJET.txt"))
##    path = codecs.decode(f"{path}",'unicode_escape')
##    path = f"{path}/{covert_PATH(P.load_element('projet'))}"
##    print(path)
    j_project =  P.load_element("Nuance")+ "\n"+P.load_element("Diameter")+"\n"+P.load_element("Epaisseur")
    
    if result_list==[]:
        print("select at leatst one interpretation ")#--------------------------------------------------------------------------
        devlabel.config(text="select at leatst one \n interpretation",bg="yellow")
        return
    if operators_names.get()=="" or operators2_names.get()=="" or POST.get()=="" or P.load_element('projet')=="" or getname=="" :
        print("fill up the info ")#--------------------------------------------------------------------------
        devlabel.config(text="Fill up the INFO enteries",bg="yellow")
        return
    if PIPE_NAME.get()!=TUBE_NAME_FROM_FILEM:
         pipe_name =PIPE_NAME.get()
         if pipe_name == "IQI":
            move_iqi_into_folder()
            print("iqi is saved")
         else:
            print("MISMATCH tube ")#------------------------------------------------------------------
            devlabel.config(text="MISMATCH TUBE NUMBER",bg="yellow")
            return
    
    if not os.path.isfile(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value)):
        
        wb = load_workbook('cn_template.xlsx')
        wb.save(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value))
        
    wb = load_workbook(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value))   
    ws = wb.active
    ws['G1'] = f"Page: {k}"
    ws['A4'] = f"Projet: {P.load_element('projet')}"
    ws['G1'].font = Font(size=18)
    ws['F4'] =  j_project
    #ws['D2'] = f'Rapport de controle\n RADIOGRAPHIE NUMIRIQUE N° "{k}"'


    # set the "Projet" and the "Post" from the form
    #Projet1 POST1 operators2_names1 operators_names1
    ws['A5'] =f'EQUIPE:   "{EQUIPE}".\nPOST:       "{POST1}".'
    #ws['A5'] =f'POST:       "{POST1}".'
    # set the name of the operators
    
    namee= operators_names1
    ws['A32'] =f'INSPECTEUR ALFAPIPE GHARDAIA:\nNom et Prenom: {operators_names1}\n Visa:'
    ws['E32'] =f'          INSPECTEUR ALFAPIPE GHARDAIA:\n          Nom et Prenom: {operators2_names1}\n          Visa:'
    
    
    
    ws['A3'] = f'DATE: {time.strftime("%d-%m-%y")}'
    ws['A3'].font = Font(size=18)
    #ws['A9'].font = Font(bold=True)
    ws['A{}'.format(j+10)] = j
    ws['B{}'.format(j+10)] = PIPE_NAME.get().upper()
    ws['C{}'.format(j+10)] = i
    
    if not var8.get():
        if not(covert_to_excel(result_list).find('(SANS HYDRO)') != -1):
            result_list.append("(SANS HYDRO)")
    if not var9.get():
        if not(covert_to_excel(result_list).find('(SANS UT)') != -1):
            result_list.append("(SANS UT)")
    if not var11.get():
        if not(covert_to_excel(result_list).find('(SANS CF)') != -1):
            result_list.append("(SANS CF)")
            

    action = "/"
    soudeur = Welder.get().upper()
    if soudeur:
        ws['D{}'.format(j+10)] = soudeur
    else:
        ws['D{}'.format(j+10)]= "/"
    
    if i ==0:
        pass
    else:
        xl_results0 =  (str(result_list)).replace('[','')
        xl_results1 =  xl_results0.replace(']','')
        xl_results2 =  xl_results1.replace("'","")
        
        if (xl_results2.find('-A-CH') != -1):
            xl_results2+=f"(CH-CN-{long}mm_{ch_def})"
            
        if soudeur=="" and (xl_results2.find('RM1') != -1):
            devlabel.config(text="ENTER LE CODE SOUDEUR",bg="yellow")
            print('enter soudeur code')
            return
        
        x = int(len(xl_results2)/50)
        if x==0:
            ws.row_dimensions[j+10].height = 20
            ws['E{}'.format(j+10)] =xl_results2
        else:
            ws.row_dimensions[j+10].height = (x+1)*20
            j_k+=x
            ws['E{}'.format(j+10)] =xl_results2

        
        if (xl_results2.find('RM1-A-REP') != -1):
            action = "A-CH ou DTP"
            

        elif (xl_results2.find('REP') != -1):
            print("reparation")
            
            if(xl_results2.find('A-CH') != -1):
                print("reparation et chute")
                action = "Rep-M 17"
                
            elif(xl_results2.find('MEULER') != -1):
                action = "Rep-M 22"

            else:
                action = "Reparation!"        
            
        elif(xl_results2.find('-A-CH') != -1):
            action = "M 17"
            print("Tube -A-CH")
            
        elif(xl_results2.find('MEULER') != -1):
            action = "M 22"
            print("Tube MEULER")
            
        elif (xl_results2.find('-FAR') != -1):
            action = "FINALE"
            
        elif(xl_results2.find('OK') != -1):
            action = "Tube OK"
            print("Tube OK")
            if not var13.get():
                copy_files_in_client_folder()

    ws[f'G{j+10}']= action
    
    #CREATE A TABLE FOR THE REPORT
    #TABLE COLUMNS
    NAME_DATABASE= "RAPPORT_CN"
    NAME_TABLE = "RAPPORT_CN"
    NAME_REPORT= "REPORT"
    col1="TUBE"
    col2="N°INTEGRATION"
    col3="CODE_SOUDEUR"
    col4="RESULTAT"
    col5="ACTION"
    col6="ZIP_IMAGES"
    
    create_atable(NAME_DATABASE,NAME_TABLE,NAME_REPORT,col1,col2,col3,col4,col5,col6)
    print("table created!")
##    NAME_TABLE =  NAME_TABLE.replace(']','_')
##    NAME_TABLE =  NAME_TABLE.replace('/','_')
##    NAME_TABLE =  NAME_TABLE.replace(' ','_')
##    NAME_TABLE =  NAME_TABLE.replace('+','_')
##    NAME_TABLE =  NAME_TABLE.replace('-','_')
    
    #INSERT A RECORD INTO THE TABLE%%%%%%%%%%%%%%%%%%%% DATABASE %%%%%%%%%%%%
    creating_the_zip_file()
    
    # RECORD DECLARATIONS
    NAME_REPORT="{}_RAPPORT_CN N°{}_{}_{}_AND_{}_{}".format(P.load_element('projet'),k,POST1,operators_names1,operators2_names1,time_value)
    REC1=PIPE_NAME.get().upper()
    REC2= i
    if soudeur:
         REC3= " SOUDEUR: "+ str(soudeur)
    else:
         REC3= ""
    REC4= xl_results2
    REC5= action
    REC6= convertToBinaryData("Images.zip")
    
    one_record=(NAME_REPORT,REC1,REC2,REC3,REC4,REC5,REC6)
    
    insert_one_record(NAME_DATABASE,NAME_TABLE,one_record)
    print("record inserted!")
    # creat an qrcode
    if var6.get() == 1 and action == "Tube OK":
        for_qr_record= covert_to_excel(str(NAME_REPORT)+" TUBE: "+str(REC1)+"_INTs: "+ str(REC2)+str(REC3)+" RESULTATs: "+ str(REC4)+" ACTION: "+ str(REC5))
        rq_genrator(for_qr_record,"QR")


    
    wb.save(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value))
    #wb._archive.close() 
    #wb.close()

    i=0
    result_list=[]
    j+=1
    DEFECT_NUMB_lab.config(text=i)
    finish_tube.config(text="Tube N°{}".format(j),bg ="GREEN2")
################################### TESTING ZONE #######################
    
################################### END TESTING ZONE #################    
    if  j>= max_line-j_k:
        
        
        ws[f'B{j+10}'] ="TOTAL"
        ws[f'B{j+10}'].font = Font(bold=True)
        try:
            ws[f'C{j+10}'] =sum([ws.cell(row=V+11, column=3).value for V in range(j-1)])
            ws.unmerge_cells(f'E{j+11}:F{j+11}')
            ws.merge_cells(f'A{j+11}:D{j+11}')
            ws.merge_cells(f'E{j+11}:G{j+11}')
            ws.row_dimensions[j+11].height = 60
            
            ws[f'A{j+11}'].font = Font(bold=True)
            #ws[f'A{j+10}'].alignment = Alignment(horizontal='left')
            ws[f'E{j+11}'].font = Font(bold=True)
            
            ws[f'A{j+11}'] =f'INSPECTEUR ALFAPIPE GHARDAIA:\nNom et Prenom: {operators_names1}\n Visa:           '
            ws[f'E{j+11}'] =f'          INSPECTEUR ALFAPIPE GHARDAIA:\n          Nom et Prenom: {operators2_names1}\n          Visa:'

            ws['A32'].value = None
            ws['E32'].value = None
            ws['B31'].value = None
            ws['C31'].value = None
        except:
            print("fill up all ")
            messagebox.showinfo("EMPTY ROWS!"," TO CALCULATE NUMBER OF INTEGRATIONS CORRECTLY,\n FILL UP ALL ROWS")
        #j+=1
        

        for bord in range(31-(j+11)):
            print("32-(j+11)===",32-(j+11))
            ws.unmerge_cells(f'E{j+12+bord}:F{j+12+bord}')
            
            for column in range(7):
                ws.cell(row=j+12+bord, column=column+1).border = NO_border
            
            


        ws.delete_rows(j+12, 30) 
        wb.save(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value))
        win32api.ShellExecute(
        0,
        "print",
        (r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value)),
        None,
        ".",
        0
        )
        #os.startfile(r"{}\POST RAPPORT N°{} at {}.xlsx".format(path,k,time_value), 'print')
        report_closedd.config(text="R-N°{}".format(k+1),bg = "green2")
        if not os.path.exists(r"{}\RAPPORTS\CN".format(path)):
            os.makedirs(r"{}\RAPPORTS\CN".format(path))
            print("the folder is createed !!")
        shutil.move(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value),r"{}\RAPPORTS\CN".format(path))
        
        finish_tube.config(text="INSERT",bg =btncolor)
        k+=1
        j=1
        result_list=[]
        j_k=1
       

        
    PIPE_NAME.config(state='normal')
    PIPE_NAME.delete(0,"end")
    Welder.delete(0,"end")
    CHUT_DISTANCE.delete(0,"end")
    CHUT_DISTANCE_check.delete(0,"end")
    list_film.selection_clear(0,END)
    print_image.deselect()
    print_image.config(bg=color)
    Prive.deselect()
    #print("name_folder===",name_folder)
    #name_folder=""
    
    
    print("!!! tube finished !!!")
    devlabel.config(text="DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2020",bg=color)
    

    
def report_closed():
    global k,i,j
##    path = covert_to_excel(load_list("PATH_PROJET.txt"))
##    path = codecs.decode(f"{path}",'unicode_escape')
##    path = f"{path}/covert_PATH({P.load_element('projet'))}"
##    print(path)

    j_project =  P.load_element("Nuance")+ "\n"+P.load_element("Diameter")+"\n"+P.load_element("Epaisseur")
    
    confirmation = messagebox.askquestion("CONFIRMATION CLOTURAGE!","VOULEZ VRAIMENT CLOTURE LE RAPPORT?")
    if not os.path.isfile(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value)):
        wb = load_workbook('cn_template.xlsx')  
        ws = wb.active
        ws['G1'] = f"Page: {k}"
        ws['G1'].font = Font(size=18)
        ws['A4'] = f"Projet: {P.load_element('projet')}"
        ws['A5'] =f'EQUIPE:   "{EQUIPE}".\nPOST:       "{POST1}".'
        namee= operators_names1
        ws['A32'] =f'INSPECTEUR ALFAPIPE GHARDAIA:\nNom et Prenom: {operators_names1}\n Visa:'
        ws['E32'] =f'          INSPECTEUR ALFAPIPE GHARDAIA:\n          Nom et Prenom: {operators2_names1}\n          Visa:'
        ws['A3'] = f'DATE: {time.strftime("%d-%m-%y")}'
        ws['A3'].font = Font(size=18)
        ws['F4'] =  j_project
        wb.save(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value))
        
    if i ==0 and confirmation=="yes":
        win32api.ShellExecute(
        0,
        "print",
        (r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value)),
        None,
        ".",
        0
        )
        
        #os.startfile(r"{}\POST RAPPORT N°{} at {}.xlsx".format(path,k,time_value), 'print')
        PIPE_NAME.delete(0,"end")
        print("!!!!!!!!rapport closed!!!!!!!!")
        report_closedd.config(text="R-N°{}".format(k+1),bg = "green2")
        finish_tube.config(text="INSERT",bg =btncolor)
        if not os.path.exists(r"{}\RAPPORTS\CN".format(path)):
            os.makedirs(r"{}\RAPPORTS\CN".format(path))
            print("the folder is created!!")
        shutil.move(r"{}\RAPPORT-CN N°{}_{}_{}.xlsx".format(path,k,POST1, time_value),r"{}\RAPPORTS\CN".format(path))
        k+=1
        i=0
        j=1
        j_k=1
        devlabel.config(text="DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2020",bg=color)
    else:
        devlabel.config(text="INSERER LE DERNIER TUBE !",bg='yellow')
        
    


def selcting_folder(event):
    global name_folder
    PIPE_NAME.delete(0,"end")
    name_folder =list_film.get(ANCHOR)
    if name_folder.startswith("A")or name_folder.startswith("B")or name_folder.startswith("C")or name_folder.startswith("D")or name_folder.startswith("E"):
        PIPE_NAME.insert(END,name_folder[:5])
    print("!!! folder selected !!!")
    


def search_button(event):
    global newpath,list_of_filmes1,getname,name, path,name_folder,path_scopy
    getname = str(PIPE_NAME.get()).upper()
    pipe_name=str(PIPE_NAME.get())
    
    if len(pipe_name)==2:
        print("length====",len(pipe_name),pipe_name[1])
        #A0001
        getname = (pipe_name[0]+"0"+"0"+"0"+pipe_name[1]).upper()
        print("getname",getname)  
    elif len(PIPE_NAME.get())==3:
        #A0012
        pipe_name=str(PIPE_NAME.get()).upper()
        #A0012
        if pipe_name == "IQI":
            getname = pipe_name
            #newpath = "{}\{}".format(path,getname)
        else:
            getname = (pipe_name[0]+"0"+"0"+pipe_name[1]+pipe_name[2]).upper()
        
    elif len(PIPE_NAME.get())==4:
        #A0123
        getname = (pipe_name[0]+"0"+pipe_name[1]+pipe_name[2]+pipe_name[3]).upper()
    elif len(PIPE_NAME.get())==5:
        #A1234
        devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2020",fg="black",bg=color)
        getname = str(PIPE_NAME.get()).upper()
        print("getname= ",getname)

        
    #if getname.startswith("A")or getname.startswith("B")or getname.startswith("C")or getname.startswith("D")or getname.startswith("E"):
       # print('only five digits are aceptable!')#---------------------------------------------------------------
       # devlabel.config(text='UNCORRECT PIPE NAME',bg="yellow")
    if (len(PIPE_NAME.get())==0):
        print('CURRENT DIRECTORIES !!')#---------------------------------------------------------------
        devlabel.config(text='CURRENT DIRECTORIES !!',bg="yellow")
        #return
##    path = covert_to_excel(load_list("PATH_PROJET.txt"))
##    path = codecs.decode(f"{path}",'unicode_escape')
##    path = f"{path}/{covert_PATH(P.load_element('projet'))}"
##    if not os.path.exists(path):
##            os.makedirs(path) 
##    print(path)
    
    if getname == "":
        list_film.delete(0,END)
        subdirs = [x[0] for x in os.walk(path)]
        last = sorted(subdirs, key=os.path.getmtime,reverse = True )
        print("the last modification is ",os.path.getmtime(path))
        print(last[-5:])
        for name_folder in last[:20]:
            folder=name_folder[-5:]
            if folder.startswith("A")or folder.startswith("B")or folder.startswith("C")or folder.startswith("D")or folder.startswith("E"):
                list_film.insert(END,folder)
        
    print("getname= ",getname)
    
    name  = str(getname[0])
    try:
        if pipe_name == "IQI":
            getname = pipe_name
            newpath = "{}\{}".format(path,getname)
        else:
            newpath = r'{}\{}\{}'.format(path,name,getname)
        os.startfile(newpath)
    except:
        pass
    
    print("newpath before * and s ===",newpath)
    
    if getname.endswith("*"):
        client_path1 = '{}\{}\{}'.format(path_client_variable,name,getname[:5])
        os.startfile(client_path1)
        client_path1=""
        
    if getname.endswith("S"):
        path_scopy1 = '{}\{}\{}'.format(path_scopy,name,getname[:5])
        os.startfile(path_scopy1)
        path_scopy1=""   
    
    PIPE_NAME.delete(0,END)
    PIPE_NAME.insert(0,getname)
    refresh_list()
    A_OK.config(bg=color)
    A_MEULER.config(bg=color)
    A_REP.config(bg=color)
    FINAL.config(bg=color)
    A_CH.config(bg=color)
    devlabel.config(text="DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2020",bg=color)
    
    
    
def refresh_list():
    global DCM_list
    list_film.delete(0,END)
    DCM_list = []
    ############
    list_of_filmes1 = os.listdir(newpath)
    for film in list_of_filmes1:
        if film.endswith(".dcm"):
            DCM_list.append(film)
            list_film.insert(END,film)

"""
    name_folder =list_film.get(ANCHOR)
    list_film.delete(0,END)
   
    if len(name_folder)>5:
        PIPE_NAME.delete(0,END)
        PIPE_NAME.insert(END,name_folder[:5])
        getname = str(PIPE_NAME.get()).upper()
        name_folder.set==""
       
    if len(name_folder)==5:
        PIPE_NAME.delete(0,END)
        PIPE_NAME.insert(END,name_folder[:5])
        getname = str(PIPE_NAME.get()).upper()
        name_folder==""

    """   
def copy_files_in_client_folder():
    global list_of_filmes,path_client_variable
    ######################################################################################################################
    #global list_of_filmes
    if var8.get() and var9.get() and var11.get():
        
        #path_client_variable = covert_to_excel(load_list("PATH_CLIENT.txt"))
        #path_client_variable = codecs.decode(f"{path_client_variable}",'unicode_escape')
        client_path = r'{}\{}\{}'.format(path_client_variable,name,getname)
        print("client path=",client_path)
          
        list_of_filmes= glob.glob(newpath+'/*.jpg')
        #shutil.make_archive(output_filename, 'zip', dir_name)
        
        #if len(list_of_filmes)==len(DCM_list):
        if not os.path.exists(client_path):
            os.makedirs(client_path) 
        
        #print("list of films is nnnnnnnnnnnnnnnnnnn=",list_of_filmes)

        for each_film in list_of_filmes:
            #print("each_film",each_film)
            if each_film.find('-A-CH') != -1 or each_film.find('MEULER') != -1 or each_film.find('RM1-A-REP') != -1 :
                pass
            #print(each_film,"not copied files__")
            else:
                shutil.copy(each_film, client_path)
                #print(each_film," copied files each file__")
                
   

def interpreted_film_func(event):
    global save_path,interpreted,ch_def,long,interpreted_film,TUBE_NAME_FROM_FILEM,newpath
    name_file =list_film.get(ANCHOR)[:-int(P.load_element("bits"))]
    name_film = name_file[6:]
    TUBE_NAME_FROM_FILEM= list_film.get(ANCHOR)[:5]
    print("interprited FILM is ",name_film)
    
    if var1.get()=="" and var2.get() == "" and var3.get()==""and var4.get()==""and var10.get()=="":
        print("select one of the decisions!")#------------------------------------------------------------------
        devlabel.config(text="SELECT A DECISION!",bg="yellow")
        return
    
    if name_file=="":
        print("select the film ")#------------------------------------------------------------------
        devlabel.config(text="SELECT ONE FILM (IMAGE)",bg="yellow")
        return
    if PIPE_NAME.get()=="":
        print("select  a tube ")#------------------------------------------------------------------
        devlabel.config(text="WRITE A TUBE NUMBER",bg="yellow")
        return
    if PIPE_NAME.get()!=TUBE_NAME_FROM_FILEM:
        pipe_name =PIPE_NAME.get()
        if pipe_name == "IQI":
            getname = pipe_name
            newpath = "{}\{}".format(path,getname)
            if var1.get()=="-OK":
                var1.set("-CONFORME")
            if var2.get()=="-A-REP":
                var2.set("-NON-CONFORME")
        else:
            print("MISMATCH tube ")#------------------------------------------------------------------
            devlabel.config(text="MISMATCH TUBE NUMBER",bg="yellow")
            return
            
    check_variable= f"{var1.get()}{var2.get()}{var3.get()}{var4.get()}{var10.get()}"
    print("check_variable",check_variable)
    save_path = f"{newpath}\{name_file}{check_variable}"
    ch_def = str(CHUT_DISTANCE_check.get())
    long = str(CHUT_DISTANCE.get())
    
    interpreted_film  = f"{name_film}{check_variable}"
    if interpreted_film:
        if (interpreted_film.find('-A-CH') != -1)and (long=="" or ch_def==""):
            print("fill up all chute informtion ")#------------------------------------------------------------------
            devlabel.config(text="FILL UP LONG AND DEFECT ENTRIES")
            return 
        
    print("---------save_path------",save_path)
    
# pyautogui application
    time.sleep(.1)
    pg.moveTo(933,87)
    pg.click()
    time.sleep(.1)
    #print_file()
    time.sleep(.1)
    pg.keyDown('ctrl')
    pg.keyDown('shift')
    pg.press('s')# SAVE
    pg.keyUp('shift')
    pg.keyUp('ctrl')
    interpreted = True
    #clicking_thread()
    #print(multiprocessing.cpu_count())
    #p1.start()
    #p1.join()
    t1.start()
    devlabel.config(text="DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2020",bg=color)
    
    
def save_film_in_specific_folder():
    global interpreted,i,result_list
    #time.sleep(1)
    if interpreted == False:
        print("interprete the filme first")#------------------------------------------------------------------
        devlabel.config(text="INSPECT FIRST BEFORE SAVING",bg="yellow")
        return
    time.sleep(.1)
    pg.moveTo(530,55)
    pg.click()
    #pg.moveTo(530,55)
    time.sleep(1)
    #pg.moveTo(100,500)
    
    pg.keyDown('ctrl')
    pg.press('s')# SAVE
    pg.keyUp('ctrl')
    time.sleep(1)
    
    if var5.get()==1:
        try:
            #the_x,the_y= pg.locateCenterOnScreen('name.png', grayscale=True,confidence = confidenceE.get())
            x, y = pg.locateCenterOnScreen('name.jpg', grayscale=True,confidence =.9)
            pg.click(x +200, y)
            print(x +200, y)
        except Exception as e:
            print("the exception is", e)
            pg.moveTo(330,855)
            pg.click()
    if var5.get()==0:
        pg.moveTo(330,855)
        #print("save coordinate= 330,855")
        #pg.moveTo(314,512)
        pg.click()

    caps_status = win32api.GetKeyState(win32con.VK_CAPITAL)
    if caps_status==0:
        print('CapsLock is off')
        file = save_path[2:]
        pg.press('capslock')
        pg.write(str(save_path[:2]))# SAVE
        pg.press('capslock')
        pg.write(file)
        print("SAVED file======== ", file)
        pg.press('capslock')
        pg.press('enter')
        
    else:
        print('CapsLock is on')
        file = save_path[2:]
        pg.write(str(save_path[:2]))# SAVE
        pg.press('capslock')
        pg.write(file)
        print("SAVED file======== ", file)
        pg.press('capslock')
        pg.press('enter')


    
    result_list.append(interpreted_film)
    i+=1
    #PIPE_NAME.config(state='disabled')
    DEFECT_NUMB_lab.config(text=i,bg="yellow")

    # rest the chek boxes
    A_REP.deselect()
    A_CH.deselect()
    A_OK.deselect()
    A_MEULER.deselect()
    FINAL.deselect()
    A_OK.config(bg=color)
    A_MEULER.config(bg=color)
    A_REP.config(bg=color)
    FINAL.config(bg=color)
    A_CH.config(bg=color)
    #print_image.deselect()
    print("anchor before= =====",list_film.get(ANCHOR))
    list_film.selection_clear(0,END)
    list_film.delete(ANCHOR)
    print("anchor after= =====",list_film.get(ANCHOR))
    
    interpreted = False
    CHUT_DISTANCE.grid_forget()
    CHUT_DISTANCE_check.grid_forget()
    CHUT_defaut_lab.grid_forget()
    CHUT_DISTANCE_lab.grid_forget()
    print("file is saved successfully !!!")
    devlabel.config(text="DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2020",bg=color)


def archiving():
    pass
##    global ARCHIVE
##    not_archived=0
##    fold_DCM_list=[]
##    fold_JPG_list_OK=[]
##    fold_JPG_list=[]
##    
##    path_client_variable = covert_to_excel(load_list("PATH_CLIENT.txt"))
##    path_client_variable = codecs.decode(f"{path_client_variable}",'unicode_escape')
##    client_path = r'{}\{}\{}'.format(path_client_variable,name,getname)
##    
##    
##    path = covert_to_excel(load_list("PATH_PROJET.txt"))
##    path = codecs.decode(f"{path}",'unicode_escape')
##    
##    arch_path  = os.path.join(path, os.pardir)
##    
##    targetA = f"{arch_path}\ARCHIVE\A"
##    targetB = f"{arch_path}\ARCHIVE\B"
##    targetC = f"{arch_path}\ARCHIVE\C"
##    targetD = f"{arch_path}\ARCHIVE\D"
##    targetE = f"{arch_path}\ARCHIVE\E"
##    
##    if not os.path.exists(targetA):
##        os.makedirs(targetA)
##    if not os.path.exists(targetB):
##        os.makedirs(targetB)
##    if not os.path.exists(targetC):
##        os.makedirs(targetC)
##    if not os.path.exists(targetD):
##        os.makedirs(targetD)
##    if not os.path.exists(targetE):
##        os.makedirs(targetE)
##    
##    subdirectoreis = [x[0] for x in os.walk(path)]
##    subdirectoreis_client = [x[0] for x in os.walk(path_client_variable)]
##    
##    ARCHIVED_FOLDERS = len(subdirectoreis)
##    for archfolder in subdirectoreis:
##        for client_folder in subdirectoreis_client:
##            print(archfolder[-5:],client_folder[-5:])
##            if archfolder[-5:]==client_folder[-5:]:
##                
##                if archfolder[-5]=="A" or archfolder[-5]=="a" :
##                    if not os.path.exists(f"{targetA}\{archfolder[-5:]}"):
##                        shutil.move(archfolder,targetA)
##                        print(archfolder,"ARCHIVED SUCCESSFULLY!!!!!!!")
##                        ARCHIVE+=1
##                        
##                elif archfolder[-5]=="B" or archfolder[-5]=="b" :
##                    if not os.path.exists(f"{targetB}\{archfolder[-5:]}"):
##                        shutil.move(archfolder,targetB)
##                        print(archfolder,"ARCHIVED SUCCESSFULLY!!!!!!!")
##                        ARCHIVE+=1
##                    
##                elif archfolder[-5]=="C" or archfolder[-5]=="c" :
##                    if not os.path.exists(f"{targetC}\{archfolder[-5:]}"):
##                        shutil.move(archfolder,targetC)
##                        print(archfolder,"ARCHIVED SUCCESSFULLY!!!!!!!")
##                        ARCHIVE+=1
##                    
##                elif archfolder[-5]=="D" or archfolder[-5]=="d" :
##                    if not os.path.exists(f"{targetD}\{archfolder[-5:]}"):
##                        shutil.move(archfolder,targetD)
##                        print(archfolder,"ARCHIVED SUCCESSFULLY!!!!!!!")
##                        ARCHIVE+=1
##                    
##                elif archfolder[-5]=="E" or archfolder[-5]=="e" :
##                    if not os.path.exists(f"{targetE}\{archfolder[-5:]}"):
##                        shutil.move(archfolder,targetE)
##                        print(archfolder,"ARCHIVED SUCCESSFULLY!!!!!!!")
##                        ARCHIVE+=1
##                    
##            #else:
##               # not_archived+=1
##                #print("not yet!!!!!!!")
##
##    archiveing_btn.config(text = f"{ARCHIVE}/{ARCHIVED_FOLDERS-ARCHIVE}",bg ="green2")
##
##
##
##


#Add the chack boxes


def check1(event):
    global var1,var2,var3,var4,var10
    if var1.get()=="-OK":
        var2.set("")
        var1.set("")
        var3.set("")
        var4.set("")
        var10.set("")
        A_OK.config(bg=color)
        A_MEULER.config(bg=color)
        A_REP.config(bg=color)
        FINAL.config(bg=color)
        A_CH.config(bg=color)
        
        CHUT_DISTANCE.grid_forget()
        CHUT_DISTANCE_check.grid_forget()
        CHUT_defaut_lab.grid_forget()
        CHUT_DISTANCE_lab.grid_forget()
        
    else:
       var1.get()=="-OK"
       var2.set("")
       var4.set("")
       var3.set("")
       var10.set("")
       A_OK.config(bg="green2")
       A_MEULER.config(bg=color)
       A_REP.config(bg=color)
       FINAL.config(bg=color)
       A_CH.config(bg=color)
       CHUT_DISTANCE.grid_forget()
       CHUT_DISTANCE_check.grid_forget()
       CHUT_defaut_lab.grid_forget()
       CHUT_DISTANCE_lab.grid_forget()
       

#var2.set(1)
def check2(event):
    global var1,var2,var3,var4,var10
    if var2.get()=="-A-REP":
        var2.set("")
        var1.set("")
        var3.set("")
        var10.set("")
        A_OK.config(bg=color)
        A_MEULER.config(bg=color)
        A_REP.config(bg=color)
        FINAL.config(bg=color)
        A_CH.config(bg=color)
        CHUT_DISTANCE.grid_forget()
        CHUT_DISTANCE_check.grid_forget()
        CHUT_defaut_lab.grid_forget()
        CHUT_DISTANCE_lab.grid_forget()
        var4.set("")
        
    else:
       var2.get()=="-A-REP"
       var1.set("")
       var3.set("")
       var10.set("")
       A_OK.config(bg=color)
       A_MEULER.config(bg=color)
       A_REP.config(bg="red")
       FINAL.config(bg=color)
       A_CH.config(bg=color)
       
       CHUT_DISTANCE.grid_forget()
       CHUT_DISTANCE_check.grid_forget()
       CHUT_defaut_lab.grid_forget()
       CHUT_DISTANCE_lab.grid_forget()
       var4.set("")

#Add the chack boxes

def check3(event):
    global var1,var2,var3,var4,var10
    if var3.get()=="-A-CH":
        var2.set("")
        var1.set("")
        var3.set("")
        var4.set("")
        var10.set("")
        A_OK.config(bg=color)
        A_MEULER.config(bg=color)
        A_REP.config(bg=color)
        FINAL.config(bg=color)
        A_CH.config(bg=color)
    else :
        var3.get()=="-A-CH"
        var1.set("")
        var2.set("")
        var4.set("")
        var10.set("")
        A_OK.config(bg=color)
        A_MEULER.config(bg=color)
        A_REP.config(bg=color)
        FINAL.config(bg=color)
        A_CH.config(bg="red2")
        CHUT_defaut_lab.grid(row = 6 , column= 0, padx = 10,sticky="W")
        CHUT_DISTANCE_check.grid(row = 7,column = 0,sticky="W",padx = 10, pady=5)
        CHUT_DISTANCE_lab.grid(row = 8 , column= 0, padx = 10,sticky="W")
        CHUT_DISTANCE.grid(row = 9 , column= 0, pady = 5, padx = 10,sticky="W")
       
        
def check4(event):
    global var1,var2,var3,var4,var10
    if var4.get()=="-A-MEULER":
        var2.set("")
        var1.set("")
        var3.set("")
        var10.set("")
        A_OK.config(bg=color)
        A_MEULER.config(bg=color)
        A_REP.config(bg=color)
        FINAL.config(bg=color)
        A_CH.config(bg=color)
        CHUT_DISTANCE.grid_forget()
        CHUT_DISTANCE_check.grid_forget()
        CHUT_defaut_lab.grid_forget()
        CHUT_DISTANCE_lab.grid_forget()
        
        
    else :
        var4.get()=="-A-MEULER"
        var1.set("")
        var2.set("")
        var3.set("")
        var10.set("")
        A_OK.config(bg=color)
        A_MEULER.config(bg="orange")
        A_REP.config(bg=color)
        FINAL.config(bg=color)
        A_CH.config(bg=color)
        CHUT_DISTANCE.grid_forget()
        CHUT_DISTANCE_check.grid_forget()
        CHUT_defaut_lab.grid_forget()
        CHUT_DISTANCE_lab.grid_forget()
        
def check10(event):
    global var1,var2,var3,var4,var10
    if var10.get()=="-FAR":
        var2.set("")
        var1.set("")
        var3.set("")
        var4.set("")
        A_OK.config(bg=color)
        A_MEULER.config(bg=color)
        A_REP.config(bg=color)
        FINAL.config(bg=color)
        A_CH.config(bg=color)
        CHUT_DISTANCE.grid_forget()
        CHUT_DISTANCE_check.grid_forget()
        CHUT_defaut_lab.grid_forget()
        CHUT_DISTANCE_lab.grid_forget()
        
        
    else :
        var10.get()=="-FAR"
        var1.set("")
        var2.set("")
        var3.set("")
        var4.set("")
        
        A_OK.config(bg=color)
        A_MEULER.config(bg=color)
        A_REP.config(bg=color)
        FINAL.config(bg="YELLOW")
        A_CH.config(bg=color)
        
        CHUT_DISTANCE.grid_forget()
        CHUT_DISTANCE_check.grid_forget()
        CHUT_defaut_lab.grid_forget()
        CHUT_DISTANCE_lab.grid_forget()

def check5(event):
    global j
    if var7.get()!=1:
        j_variable_lab.grid(row = 7 , column= 0, padx = 10,pady=10,sticky="E")
        k_variable_lab.grid(row = 8 , column= 0, padx = 10,sticky="E")
        j_variable.grid(row = 7 , column= 1, padx = 10,pady= 10,sticky="W")
        k_variable.grid(row = 8 , column= 1, padx = 10,pady=10,sticky="W")
        j_variable_btn.grid(row =8,column=1, padx = 10,pady=10,sticky="E")
        #j = int(j_variable.get())
        root.geometry("370x345")
        #print("var7=", var7.get(),"j=",j)
                
    else:
        j_variable.grid_forget()
        k_variable.grid_forget()
        j_variable_lab.grid_forget()
        k_variable_lab.grid_forget()
        root.geometry("370x250")
        j_variable_btn.grid_forget()
        #print("var7=", var7.get(),"j=",j)

app_state = False       

def show_frame(frame):
    global path,path_client_variable,path_scopy,app_state
    
    if str(operators_names.get())=="" or  str(operators2_names.get())=="" or str(POST.get())=="" or str(P.load_element('projet'))=="" :
        operator1_names_lab.config(bg = "orange1")
        operator2_names_lab.config(bg = "orange1")
        POST_lab.config(bg = "orange1")
        Projet_lab.config(bg = "orange1")
        #EQUIPE_lab.config(bg = "orange1")
        return
    
    
    frame.tkraise()
    if frame ==PIPE_FRAME:
        
        
        root.geometry("410x900+500+90")
        try:
            
            #print("this is the password you wrot =",f"{operators_names.get()} ",f"{PASSWORD_OP.get()}")
            if not app_state:
                #app = application.Application().start(r"C:\Windows\System32\mspaint.exe")
                app = application.Application().start(P.load_element("PATH_prog"))#r"C:\Program Files\YXLON\Y.Image 3500\IMAGE_3500.exe")
                app_state = True
    
            time.sleep(.5)
            app.IMAGE3500Etablissementdelaconnexion.ComboBox.Select(f"{operators_names.get()}")
            app.IMAGE3500Etablissementdelaconnexion["enter votre mot de passe:Edit"].type_keys(f"{PASSWORD_OP.get()}")
            
            #arret
            #app.IMAGE3500Etablissementdelaconnexion.Button2.click()***
            #ok
            app.IMAGE3500Etablissementdelaconnexion.Ok.click()
            #app.IMAGE3500DD.Ok.click()
            PASSWORD_OP.delete(0,"end")
        except Exception as e :print(e,"no programe is found")
        
        path = P.load_element("PATH_PROJET")
        
        path_client_variable = P.load_element("PATH_CLIENT")
        #path_client_variable = codecs.decode(r"{}".format(path_client_variable),'unicode_escape')
        path_scopy = P.load_element("PATH_SCOPIE")
        
        #covert_to_excel(load_list("PATH_PROJET.txt"))
        #path = codecs.decode(f"{path}",'unicode_escape')
        
        path = f"{path}\{covert_PATH(P.load_element('projet'))}"
        print("path---***----",path)
        
        if not os.path.exists(path):
                os.makedirs(path) 
        print(path)
        
    if frame ==starting:
        root.geometry("410x280+500+90")
        
    frame.grid(row =0,column=0,sticky='nsew')
    
def j_variablefunc():
    global j,k
    if var7.get()==1:
        j = int(j_variable.get())
        k = int(k_variable.get())
        j_variable_btn.config(text=f"LN°{j}/RN°{k}",bg="green2")
        
        print("var7=", var7.get(),"j=",j)

##########################################################################################################
def quit_func():
    global quit_var
    quit_var = True
    print("quit var=",quit_var)
    root.destroy()
    

    
def clicking_thread():
    global a
    while True:       
        a = win32api.GetKeyState(0x02)
       # print("a===",a)
        if (interpreted == True and a<0):
            print("the right key is clicked", a)
            save_film_in_specific_folder()
        if quit_var:
            break
        



print("thread one and 3 are statted")
#import multiprocessing




# 

#########################################################################################################################
root = Tk()
root.config(bg ="white")
root.focus_force()
#color = "cyan2"
color = "light sky blue"
color = "deep sky blue"
btncolor ="gold"
root.rowconfigure(0,weight= 1)
root.columnconfigure(0,weight=1)

#Add the chack boxes
var1 = StringVar()       
var2 = StringVar()
var3 = StringVar()
var4 = StringVar()
var5 = IntVar()
var6 = IntVar()
var7 = IntVar()
var8 = IntVar()
var9 = IntVar()
var11 = IntVar()
var12 = BooleanVar()
var13 = BooleanVar()
var10= StringVar()
distance = 7
dis_fon=13



starting = tk.Frame(root, width=100, height=100, background=color)
starting.grid(row =0,column=0,sticky='nsew')

#PATH_FRAME = tk.Frame(root, width=100, height=100, background=color)
#PATH_FRAME.pack(fill= "both", expand ="YES" , pady = 0, padx=10)

PIPE_FRAME = tk.Frame(root , width=50, height=50, background=color)
#PIPE_FRAME.pack(fill= "both", expand ="YES" , pady = 0, padx=10)

decision_FRAME = LabelFrame(PIPE_FRAME, text = "DECISION",  width = 35,height =20,font =("Helvetica",15,"bold"),bg =color )
decision_FRAME.grid(row = 1 , column= 2, pady = 10,sticky="N",rowspan= 4)



starting_FRAME = LabelFrame(starting, text = "INFO",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color )
starting_FRAME.pack(pady = 10,anchor="center")



############################################################----FARME STARTING----- ####################################################################################################################################################################################################


#OPERATOR1 LABEL
operator1_names_lab = Label(starting_FRAME, text= "INSPECTEUR 01:",font =("Helvetica",10,"bold"), bg =color)
operator1_names_lab.grid(row = 0 , column= 0, padx = 10,sticky="W", pady=5)


# combobox5
def operators_namesfunc(event):
    global operators_names1
    operators_names1  = operators_names.get()
    operator1_names_lab.config(bg="green2")
    PASSWORD_OP.focus_set()
    
    print(operators_names1)

## Adding combobox FILM A REFAIR
n4 = tk.StringVar() 
 
operators_names = ttk.Combobox(starting_FRAME, width = 15, textvariable = n4,font= ("Courier", 12, "bold"),stat="readonly" )

operators_names['values'] =P.load_element("LISTE_operateurs")#load_list("LISTE_operateurs.txt")
  
operators_names.grid(column = 1, row = 0,sticky="W",padx = 10, pady=5) 
operators_names.current()
operators_names.bind("<<ComboboxSelected>>", operators_namesfunc)
operators_names.focus_set()



PASSWORD_OP_lab = Label(starting_FRAME, text= "PASSWORD:",font =("Helvetica",10,"bold"), bg =color)
PASSWORD_OP_lab.grid(row = 1 , column= 0, padx = 10,sticky="W", pady=5)

def password_func(event):
    operators2_names.focus_set()
    

PASSWORD_OP = Entry(starting_FRAME, width = 18,relief ="sunken", font =("Helvetica",12), bg ="white",show = '*')
PASSWORD_OP.grid(row = 1 , column= 1,padx = 10, pady = 1,columnspan=2,sticky="W")
PASSWORD_OP.bind('<Return>', password_func)

 
operator2_names_lab = Label(starting_FRAME, text= "INSPECTEUR 02:",font =("Helvetica",10,"bold"), bg =color)
operator2_names_lab.grid(row = 2 , column= 0, padx = 10,sticky="W", pady=5)


# combobox5
def operators2_namesfunc(event):
    global operators2_names1
    operators2_names1  = operators2_names.get()
    operator2_names_lab.config(bg="green2")
    print(operators2_names1)

## Adding combobox FILM A REFAIR
n4 = tk.StringVar() 
 
operators2_names = ttk.Combobox(starting_FRAME, width = 15, textvariable = n4,font= ("Courier", 12, "bold"),stat="readonly" )

operators2_names['values'] =P.load_element("LISTE_operateurs")#load_list("LISTE_operateurs.txt")
  
operators2_names.grid(column = 1, row = 2,sticky="W",padx = 10, pady=5) 
operators2_names.current()

operators2_names.bind("<<ComboboxSelected>>", operators2_namesfunc)



# POST LABEL

POST_lab = Label(starting_FRAME, text= "POST:",font =("Helvetica",10,"bold"), bg =color)
POST_lab.grid(row = 4 , column= 0, padx = 10,sticky="W", pady=5)

list_des_post=P.load_element("LISTE_DES_POST")#load_list("LISTE_DES_POST.txt")
def POSTfunc(event):
    global POST1,EQUIPE
    #POST1  = POST.get()[1:]
    timing = datetime.now()
    if int(list_des_post[0])<=timing.hour<int(list_des_post[1]) :
        POST1="1 er"
    if int(list_des_post[2])<=timing.hour<int(list_des_post[3]) :
        POST1="2 eme"
    if int(list_des_post[4])<=timing.hour<int(24) or int(0)<=timing.hour <int(list_des_post[5]):
        POST1="3 eme"
        #print("post------*******-------",POST1)
        
    
    
    EQUIPE= POST.get()[0]
    POST_lab.config(bg="green2")
    print("post=",POST1,"equipe=",EQUIPE)


    

n5 = tk.StringVar() 
 
POST = ttk.Combobox(starting_FRAME, width = 8, textvariable = n5,font= ("Courier", 12, "bold"),stat="readonly" )

POST['values'] =('A',  
                 'B', 
                 'C',
                 'D',  
                 )
  
POST.grid( row = 4,column = 1,sticky="W",padx = 10, pady=5) 
POST.current()
POST.bind("<<ComboboxSelected>>", POSTfunc)



# Projet

# Projet LABEL

##Projet_lab = Label(starting_FRAME, text= "PROJET:",font =("Helvetica",10,"bold"), bg =color)
##Projet_lab.grid(row = 5 , column= 0, padx = 10,sticky="W", pady=5)
##
##
##def Projetfunc(event):
##    global Projet1
##    Projet1  = P.load_element('projet')
##    Projet_lab.config(bg="green2")
##    print(Projet1)
##
##
##
##n5 = tk.StringVar() 
## 
##Projet = ttk.Combobox(starting_FRAME, width = 8, textvariable = n5,font= ("Courier", 12, "bold"),stat="readonly" )
###Projet['values'] =P.load_element("LISTE_Projets")
###Projet['values'] =load_list("LISTE_Projets.txt")
##  
###Projet.grid(row = 5,column = 1,sticky="W",padx = 10, pady=5) 
##Projet.current()
##Projet.bind("<<ComboboxSelected>>", Projetfunc)



fr1_btn= tk.Button(starting_FRAME,text="ENTER",bg = "GOLD",command=lambda:show_frame(PIPE_FRAME),font =("Helvetica",13,"bold"),height = 1, width = 10)
fr1_btn.grid(row = 6 , column= 1,padx = 10, pady = distance,sticky="E")


continueing_checkbtn= Checkbutton(starting_FRAME, text = "RAPPORT INCOMPLET",font =("Helvetica",8,"bold"), variable = var7, bg =color)
continueing_checkbtn.grid(row = 6 , column= 0,padx = 10, pady = distance,sticky="W")
continueing_checkbtn.bind('<Button-1>',check5)


j_variable_lab = Label(starting_FRAME, text= "LineN°:",font =("Helvetica",10,"bold"), bg =color)
k_variable_lab = Label(starting_FRAME, text= "Rapport N°:",font =("Helvetica",10,"bold"), bg =color)
j_variable = Spinbox(starting_FRAME,from_=1, to = (max_line-1) ,bg ="white",increment =1,width = 3, font =("Helvetica",10),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
k_variable = Spinbox(starting_FRAME,from_=1, to = max_line ,bg ="white",increment =1,width = 3, font =("Helvetica",10),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")

#j_variable.grid(row = 4 , column= 1, pady = 10,columnspan = 2)
#j_variable.delete(0,"end")
#j_variable.insert(0,"8")
#j_variable.bind("<<SpinboxSelected>>",j_variablefunc)

j_variable_btn= tk.Button(starting_FRAME,text="VALIDE",bg = "GOLD",width = 10,command=j_variablefunc ,font =("Helvetica",13,"bold"),height = 1)

    



###########################################################----END OF FARME STARTING----- ###########################################################################################################################################################################


############################################################----FARME PIPE_FRAME----- ####################################################################################################################################################################################################

def deleting_in_pipe_entry(event):
    if i!=0:
        confirmation_insertion= messagebox.askquestion("CONFIRMATION TUBE!","VOULEZ VRAIMENT SUPREMER LE TUBE COURANT?")
        if confirmation_insertion=="yes":
            pass
            #PIPE_NAME.delete(0,"end")
        else:
            print("confirmation refused!!")
            pass
    else:
       PIPE_NAME.delete(0,"end")
       


# PIPE_FRAME WEDGITS

PIPE_NAME_lab = Label(PIPE_FRAME, text= "PIPE: ",font =("Helvetica",10,"bold"), bg =color)
PIPE_NAME_lab.grid(row = 0 , column= 0, padx = 10,pady = 10, sticky="W")

# pipe name 
PIPE_NAME = Entry(PIPE_FRAME, width = 13,relief ="sunken", font =("Helvetica",15), bg ="white")
PIPE_NAME.grid(row = 0 , column= 0,padx = 15, pady = 10,columnspan=2,sticky="E")
PIPE_NAME.focus_set()
PIPE_NAME.bind('<Return>',search_button)
#PIPE_NAME.bind('<Button-1>',search_button)
#PIPE_NAME.bind('<BackSpace>',deleting_in_pipe_entry)



btn_saerch=Button(PIPE_FRAME,text ="OUVRE", bg =btncolor,activebackground="BLUE1",font =("Helvetica",13,"bold"),height = 2, width = 10)
btn_saerch.grid(row = 0 , column= 2, padx = 15,pady = 10,sticky="E")
#btn_saerch.focus_set()
btn_saerch.bind('<Return>',search_button)
btn_saerch.bind('<Button-1>',search_button)



list_film = Listbox(PIPE_FRAME,width = 25,height =21,font =("Helvetica",12,"bold"), bg ="white")
list_film.grid(row = 1 , column= 0, padx = 10,pady = 15,rowspan =5, columnspan=2, sticky="E")
#list_film.bind("<Button-1>",selcting_folder)
list_film.bind("<Double-Button>",selcting_folder)

#refresh_button=Button(PIPE_FRAME,text ="Refresh", bg =btncolor,activebackground="GRAY1",font =("Helvetica",10,"bold"),height = 2, width = 21,command = refresh_list)
#refresh_button.grid(row = 2 , column= 1, padx = 10,sticky="W")

Welder_lab = Label(PIPE_FRAME, text= "SOUDEUR: ",font =("Helvetica",10,"bold"), bg =color)
Welder_lab.grid(row = 7 , column= 0, padx = 10,pady = 10)

# pipe name 
Welder = Entry(PIPE_FRAME, width = 5,relief ="sunken", font =("Helvetica",15), bg ="white")
Welder.grid(row = 7 , column= 1,padx = 15, pady = 10, sticky="W")





DEFECT_NUMB_lab_lab = Label(PIPE_FRAME, text= "INT/TUBE:",font =("Helvetica",10,"bold"), bg =color)
DEFECT_NUMB_lab_lab.grid(row = 8 , column= 0, padx = 10, pady=10)

DEFECT_NUMB_lab = Label(PIPE_FRAME, text= i,font =("Helvetica",10,"bold"), bg =color)
DEFECT_NUMB_lab.grid(row = 8 , column= 1, padx = 10,sticky="W")

use_saerch_image = Checkbutton(PIPE_FRAME, text = "AUTO", variable = var5, bg =color)
#use_saerch_image.grid(row = 7, column= 2,padx = 5, pady =10)
use_saerch_image.select()

SCAN_btn=Button(PIPE_FRAME,text ="AI.SCAN",bg =btncolor,activebackground="GREEN2",font =("Helvetica",dis_fon,"bold"),height = 2, width = 10)
SCAN_btn.grid(row = 7 , column= 2, padx = 5,pady = distance)
SCAN_btn.bind('<Return>',read_dicom)
SCAN_btn.bind('<Button-1>',read_dicom)
set_value=StringVar()
set_value.set("10")

scan_slider = Spinbox(PIPE_FRAME,from_=-1000, to = 1000 ,bg ="white",increment =1,width = 3, textvariable=set_value,xscrollcommand=True ,font =("Helvetica",dis_fon),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
scan_slider.grid(row = 8 , column= 2, padx = 10)
scan_slider.bind('<Button-1>',slider_update)
scan_slider.bind('<Return>',slider_update)
scan_slider.bind('<MouseWheel>',adding_numbers_toslider)



# NOMBER D'INTEGRATION LABEL

# hydro and UT verification
HYDRO_STAT = Checkbutton(PIPE_FRAME, text = "HYDRO", onvalue=1, offvalue=0, variable = var8, bg =color)
HYDRO_STAT.grid(row = 9, column= 0,padx = 5, pady =10)
HYDRO_STAT.select()

UT_STAT = Checkbutton(PIPE_FRAME, text = "UT", onvalue=1, offvalue=0, variable = var9, bg =color)
UT_STAT.grid(row = 9, column= 1,padx = 5, pady =10)
UT_STAT.select()

CF_STAT = Checkbutton(PIPE_FRAME, text = "CF", onvalue=1, offvalue=0, variable = var11, bg =color)
CF_STAT.grid(row = 9, column= 2,padx = 5, pady =10)
CF_STAT.select()

LIVE_SCOPY = Checkbutton(PIPE_FRAME, text = "SCOPY", variable = var12, bg =color)
LIVE_SCOPY.grid(row = 10, column= 0,padx = 0, pady =10)
#LIVE_SCOPY.select()
LIVE_SCOPY.bind('<Button-1>',check_live)


Prive = Checkbutton(PIPE_FRAME, text = " ", variable = var13, bg =color)
Prive.grid(row = 10, column= 2,padx = 0, pady =10)
#Prive.select()


v=1
def check(event):
    global v
    if var6.get()==0:
        print_image.config(bg ="orange")
    elif var6.get()==1:
        print_image.config(bg=color)
        
print_image = Checkbutton(PIPE_FRAME, text = "QR", variable = var6, bg =color)
print_image.grid(row = 12 , column= 1,padx = 5, pady = 10)
print_image.bind('<Button-1>',check)
print("v6============",var6.get())



finish_tube=Button(PIPE_FRAME,text ="INSERE", bg =btncolor,activebackground="YELLOW",font =("Helvetica",13,"bold"),height = 2, width = 10,command = tube_finished)
finish_tube.grid(row =12, column= 0, padx = 10,sticky="W", pady=10)
#finish_tube.bind('<Return>',tube_finished)
#finish_tube.bind('<Button-1>',tube_finished)

report_closedd=Button(PIPE_FRAME,text ="CLOTURE", bg =btncolor,activebackground="red",font =("Helvetica",13,"bold"),height = 2, width = 10,command = report_closed)
report_closedd.grid(row = 12 , column= 2,sticky="E", pady=10)
#report_closedd.bind('<Return>',report_closed)
#report_closedd.bind('<Button-1>',report_closed)


archiveing_btn=Button(PIPE_FRAME,text ="ARCHIVE", bg =btncolor,activebackground="BLUE",font =("Helvetica",13,"bold"),height = 2, width = 10,command = archiving)
#archiveing_btn.grid(row =10, column= 2, sticky="E", pady=10,padx = 20)


#quit_btn= tk.Button(PIPE_FRAME,text=" QUIT  ",bg =btncolor,activebackground="RED",font =("Helvetica",10,"bold"),height = 2,width = 11,command=quit_func)
#quit_btn.grid(row = 9, column= 2, padx =15,sticky="W", pady=10)

fr3_btn= tk.Button(PIPE_FRAME,text="RETOUR>>", width = 10,bg =btncolor,activebackground="cadetblue1",command=lambda:show_frame(starting))
fr3_btn.grid(row = 18 , column= 1,padx = 10, pady =5,sticky="E")

devlabel = Label(PIPE_FRAME, text= "BOUZID CND-RT-II 2020",font =("Algerian",10,"bold"), bg =color)
devlabel.grid(row = 17 , column= 0, columnspan = 3, pady = 10, padx= 60)


############################################################----END OF FARME PIPE_FRAME----- ####################################################################################################################################################################################################


################################################################-----DECISION FRAME------############################################################################


A_OK= Checkbutton(decision_FRAME, text = "OK", variable = var1,onvalue="-OK",offvalue="", bg =color, font =("Helvetica",dis_fon))
A_OK.grid(row = 1 , column= 0,padx = 10, pady = distance,sticky="W")
A_OK.bind('<Button-1>',check1)

A_MEULER= Checkbutton(decision_FRAME, text = "A MEULER", variable = var4,onvalue="-A-MEULER",offvalue="", bg =color, font =("Helvetica",dis_fon))
A_MEULER.grid(row = 2 , column= 0,padx = 10, pady = distance,sticky="W")
A_MEULER.bind('<Button-1>',check4)


A_REP= Checkbutton(decision_FRAME, text = "A REP", variable = var2, onvalue="-A-REP",offvalue="", bg =color, font =("Helvetica",dis_fon))
A_REP.grid(row = 3 , column= 0,padx = 10, pady = distance,sticky="W")
A_REP.bind('<Button-1>',check2)

FINAL= Checkbutton(decision_FRAME, text = "FAR/FAP", variable = var10, onvalue="-FAR",offvalue="", bg =color, font =("Helvetica",dis_fon))
FINAL.grid(row = 4 , column= 0,padx = 10, pady = distance,sticky="W")
FINAL.bind('<Button-1>',check10)
        
A_CH= Checkbutton(decision_FRAME, text = "A CHUTE", variable = var3,onvalue="-A-CH",offvalue="", bg =color, font =("Helvetica",dis_fon))
A_CH.grid(row = 5 , column= 0,padx = 10, pady = distance,sticky="W")
A_CH.bind('<Button-1>',check3)


CHUT_DISTANCE = Entry(decision_FRAME, width = 5,relief ="sunken", font =("Helvetica",dis_fon), bg ="white")
CHUT_DISTANCE_lab = Label(decision_FRAME, text= "Long (mm):",font =("Helvetica",dis_fon,"bold"), bg =color)
CHUT_defaut_lab = Label(decision_FRAME, text= "Defaut:",font =("Helvetica",dis_fon,"bold"), bg =color)


def CHUT_DISTANCE_checkfunc(event):
    global CHUT_DISTANCE_check1
    CHUT_DISTANCE_check1  = CHUT_DISTANCE_check.get()
    print(CHUT_DISTANCE_check1)

def CHUT_DISTANCE_checkdel(event):
    global CHUT_DISTANCE_check1
    CHUT_DISTANCE_check1  = ""
    print("Projet1 is deleted!")


n7 = tk.StringVar() 
 
CHUT_DISTANCE_check = ttk.Combobox(decision_FRAME, width = 5, textvariable = n7,font= ("Courier", dis_fon, "bold"),stat="readonly" )

CHUT_DISTANCE_check['values'] =P.load_element("DEFAULT_ACH")#load_list("DEFAULT_ACH.txt")
  

CHUT_DISTANCE_check.current()
CHUT_DISTANCE_check.bind("<<ComboboxSelected>>", CHUT_DISTANCE_checkfunc)
CHUT_DISTANCE_check.bind("<BackSpace>", CHUT_DISTANCE_checkdel)



Interpreter_btn=Button(decision_FRAME,text ="INSPECTE",bg =btncolor,activebackground="GREEN2",font =("Helvetica",dis_fon,"bold"),height = 2, width = 10)
Interpreter_btn.grid(row = 100 , column= 0, padx = 5,pady = distance)
Interpreter_btn.bind('<Return>',interpreted_film_func)
Interpreter_btn.bind('<Button-1>',interpreted_film_func)




###########################################################------END OF DECISION FRAME-------###################################################################################
###########################################################-----------THREADS---------###################################################################################

t1=Thread(target=clicking_thread)
t1.start
#run=run_stram()
t3=Thread(target=RUN_stream)
#t3.start()
#if not t3.is_alive():
        


#OPERATOR2 LABEL####################################################################################################################################################

def doSomething():
    global quit_var
    quit_var = True
    quitt()
    print("quit var=",quit_var)
    root.destroy()
    

icon = """AAABAAEA4eEAAAEACACs6AAAFgAAACgAAADhAAAAwgEAAAEACAAAAAAAZMgAAAAAAAAAAAAAAAEAAAABAAA4/f8AGxYZAP///wDu7u4A7e3tADn//wAAAAAA+fn5APT09AD7+/sAFwAAADCvsQAVAAAAGAAAABsVGAATAAAAJ4mKAC+VlgAaERQAGQoOABYQFAAbAgkAztjZAAoAAAAbAAAAycjJABkRFAAfbW4AABcZABkGCwAPAAAAGAwPADjx8gAfAAAANcnJADPW1wAWFxkAI0ZIADfw8QDY3t8Ay9DQAGBdXQBNSUoAPDY3AB0kJgBrb3AAABEVANrZ2QB9enoAH3R2ABpMTQBUWlsAuLe3ADJydAAyvL4AK01PADk9PgAjW1wAJDEzADF/gQCPlJUAI2VnAJmXlwA35+gAIxEWADWztABqZ2gAp6usAGt8fgClo6MAN0lLADxGSAAmHR4Asby9AJCVlgCRoaIAJywtAKCvsQAbMjQALcDBADpPUQAyeXoAWGttAC2iowAtZWYAGz5AAIWFhgAAHyIAc46QABYqKwA7q6wAFh4gADaTlAAuPD4ALi0vADFeYAAwVFYAM2BiAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwgHBwkJAgICBAQnFhYWKCgWFhYWBAQEAgICAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQkIAwQ0NFhEQikpKytMHBwcBgYuHBwcKysrKSktREM0NAMIBwICCAgIBAQEAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICBElJRTBXBgYGBgYGFxQUFBQUEhISDg4OEhIaFBQUFBcXBgYGBgYGVzBFSUkIAgIJAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMEAwgHAgQnNFY4XgYGBhcSDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaEhcGBgZIRzwvJwQCAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMCAgJNVjgGBgYUFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEUHgYGBjhWGQICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAwcJAi9FLAYGBhQOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhQGBgYsRSgCAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBwREKRQGGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaBhQpRAQJCQgDBAMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAi8wKgYGFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEUBgYqRS8CCQMEAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEBAICAgICAgICAgcEBAQEBAQEBAMCAgICAgICAgIIBAQEBAQEBAQIAgICAgRFRh8GFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEUBh9QRQQCCAQEBwICAgICAgICCQQEBAQEBAQEBAICAgICAgICAgcEBAQEBAQEBAMCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAABwcHBwcHBwcHCAgICAgICAgICAcHBwcHBwcHBwgICAgICAgICAcHBwcHBwcHBwcICAgICAgICAgHBwcHBwcHBwcHAgIZVl4GHgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhITFRUKDQ0MDAoNDQwMGBUVExIOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4eBl5KAwICCAgICAgICAgICAcHBwcHBwcHBwgICAgICAgICAcHBwcHBwcHBwcICAgICAgICAgHBwcHBwcHBwcHCAgICAgICAgIAAAAAgICAgICAgICCAQDAwMDAwMECAICAgICAgICAgMEAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIIGTgGBhQBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4fHQoMDA8MEzoyGxsQEBAQEBAQEBAQEBsbMjpADwoMDAodEg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBARQGBkQZAgcDBAQDAwMECAICAgICAgICAgMEAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICBxk4Bh4BAQEBAQEBAQEBAQEBAQEBAQEBAQEOGh0NDB9ZJSU1Cz8/Pz8gAAAFBQUFBQUFBQUFBQAAID8/IyILNSUlWR8MDR0aAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgYGMy8JCQMEAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgJDTgYeAQEBAQEBAQEBAQEBAQEBAQEBAQ4TDQ8PQDlTQSIABQUFBQUFBQUFBQUAAAAAAAAAAAAAAAUFBQUFBQUFBQUFACJBETlADA8NHw4BAQEBAQEBAQEBAQEBAQEBAQEBFwYpGQIJBAMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgIISToGHwEBAQEBAQEBAQEBAQEBAQEBAQEfDRQyPUEmIAAFBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFBQAgJhE9VRQKHwEBAQEBAQEBAQEBAQEBAQEBAQEXBioZCQcECAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQCAhlHBhQBAQEBAQEBAQEBAQEBAQEBAQ4NDwwsCyMgBQUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUFPyMLLB4eDQ4BAQEBAQEBAQEBAQEBAQEBAR4GQgQCCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMECAIWOAYeAQEBAQEBAQEBAQEBAQEBAQ4dDyExIgAFBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFBQVaMSEPDBIBAQEBAQEBAQEBAQEBAQEBFwYtJwICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDBAMCCUQsFwEBAQEBAQEBAQEBAQEBAQETDRQ3Nj8FBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBT82NxQKEgEBAQEBAQEBAQEBAQEBAQEGLD4CAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQDAwMDAwMECAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMEBwJDDgYUAQEBAQEBAQEBAQEBAQEaDw9VCyMFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBSMLIQ8NDgEBAQEBAQEBAQEBAQEBGgYqSQICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAACQkJCQkJCQkJCAgICAgICAgICAkJCQkJCQkJCQgICAgICAgICAcJCQkJCQkJCQkICAgCBEIGFAEBAQEBAQEBAQEBAQEBGg8SUU8ABQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFADY5DAoOAQEBAQEBAQEBAQEBAQEXBkIEAgkJCQgICAgICAgICAcJCQkJCQkJCQkICAgICAgICAgHCQkJCQkJCQkHCAgICAgICAgIAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEBAICAgICAgICAgcEBAQEBAQEBAMCAgJLXgYBAQEBAQEBAQEBAQEBAR8PDREjBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQUFBQUFBQUFBQUFBQUFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFIlQeDxoBAQEBAQEBAQEBAQEBDgZdNAIIBAICAgICAgICAgcEBAQEBAQEBAMCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDBAMCLzMGFAEBAQEBAQEBAQEBAQETDFURAAUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQUFBQUFBQUAJj8jIiIiIiIiIiIjIz8mAAUFBQUFBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQU/EFsMEgEBAQEBAQEBAQEBAQEUBkQECQICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMECAlJXgYBAQEBAQEBAQEBAQEBEgxOECAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQUFACAgP1M7PT05Mk4SEwwMGA0NDRgOEk4yOT09XAs/ICAABQUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBT9UWwwOAQEBAQEBAQEBAQEBAQZeKAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwQHCC0GFAEBAQEBAQEBAQEBARoMDBEFBQUAAAAAAAAAAAAAAAAAAAAAAAAABQUFBQUgIwsxXQ0KDw8PDRUdExIODgEBDg4ODg4SDhITHRUNDw8MDQ0lUwsjIAUFBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAUFJjsYDA4BAQEBAQEBAQEBAQEUBkUJAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwIvWQYBAQEBAQEBAQEBAQEBCh0QIAUFAAAAAAAAAAAAAAAAAAAAAAAAAAUFAD8/OyVZFA0NHw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4TDRgkWSU7PyYABQAAAAAAAAAAAAAAAAAAAAAAAAAABQUmNQwTAQEBAQEBAQEBAQEBDgZQCAICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAksPHgEBAQEBAQEBAQEBARMeMiIFBQAAAAAAAAAAAAAAAAAAAAAAAAUFBSBTG0gMDBMSDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEOEhMMDFUbUyAFBQUAAAAAAAAAAAAAAAAAAAAAAAUFBTZbDxIBAQEBAQEBAQEBAQEXXhkCAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEBAgCRAYUAQEBAQEBAQEBAQEOHk5BBQUAAAAAAAAAAAAAAAAAAAAABQUFBSMRQB4PFRoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR8VDx5AESMFBQUFAAAAAAAAAAAAAAAAAAAAAAUFXFsMDgEBAQEBAQEBAQEBHwZYAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcECQRCBg4BAQEBAQEBAQEBAQoNED8FAAAAAAAAAAAAAAAAAAAAAAUFBQARPSQKHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfCiQ9EQAFBQAAAAAAAAAAAAAAAAAAAAAABT89DB8BAQEBAQEBAQEBAQ4GPAICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEBAICAgICAgICAggCJysGAQEBAQEBAQEBAQEaD1U2BQUAAAAAAAAAAAAAAAAAAAAFBQVPMSEPFRoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoVDyExTwUFBQAAAAAAAAAAAAAAAAAAAAUFCyQKDgEBAQEBAQEBAQEOBkIDAggEBAQEBAQEBAMCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDAwIoWR4BAQEBAQEBAQEBAR8PNQUFAAAAAAAAAAAAAAAAAAAAAAUgCzIPDRoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoNDFVTJgUFAAAAAAAAAAAAAAAAAAAABT85DxoBAQEBAQEBAQEBAQYqJwICAgICAgICAgIDAwMDAwMDAwMJAgICAgICAgIHAwMDAwMDAwMDAAAAAgICAgICAgICCAQDAwMDAwMDCAICAgICAgICAgMDAwMDAwMEAigGHwEBAQEBAQEBAQEBDyEiBQUAAAAAAAAAAAAAAAAABQUFBVpVDAoOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhMMThEmBQUAAAAAAAAAAAAAAAAAAAUFUwwdAQEBAQEBAQEBAQEXVycCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwQCNAYaAQEBAQEBAQEBAQ4MPSAFBQAAAAAAAAAAAAAAAAAFBSNTCg8TAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEg8KECMFBQAAAAAAAAAAAAAAAAAFBT8lDwEBAQEBAQEBAQEBHwYoAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDCAJWBgEBAQEBAQEBAQEBEx4RAAUAAAAAAAAAAAAAAAAABQUgUyQKGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaCiRTIAUFAAAAAAAAAAAAAAAAAAU/Gw0SAQEBAQEBAQEBARQGNAICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAj4GAQEBAQEBAQEBAQEKJE8FBQAAAAAAAAAAAAAAAAAFBVMVDxoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoPFVMABQAAAAAAAAAAAAAAAAAFBVMeEwEBAQEBAQEBAQEaBhkCAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwgHVgYBAQEBAQEBAQEBAQ8lJgUAAAAAAAAAAAAAAAAAAAAFIw0fAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHRMgBQAAAAAAAAAAAAAAAAAAAAUjWx0BAQEBAQEBAQEBDgY0AgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDBAI+BhQBAQEBAQEBAQEBHkYgBQAAAAAAAAAAAAAAAAAAAAAFBTsMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDVoFAAAAAAAAAAAAAAAAAAAAAAUFPyQMAQEBAQEBAQEBARQGGQICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMEAlYGAQEBAQEBAQEBARIPEQUFAAAAAAAAAAAAAAAAAAAAAAAABSYkHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKOQUAAAAAAAAAAAAAAAAAAAAAAAAABQBfDw4BAQEBAQEBAQEOBjQCAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAgMCVgYBAQEBAQEBAQEBGgwRBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAUQDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4KIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAUgGw8OAQEBAQEBAQEBGgYoAgIEBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAABwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHCAIZBgEBAQEBAQEBAQESDDYFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgVRMBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwbBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAFABEPDgEBAQEBAQEBARQGJwIIBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEAhkGDgEBAQEBAQEBAR8fNgUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFNg8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhgmBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQVTDw4BAQEBAQEBAQEfVycCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwQJJy4aAQEBAQEBAQEBDg82BQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSUNAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFEA8BAQEBAQEBAQEBFyoDAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwgEKxcBAQEBAQEBAQESDyMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMkEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdMiAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBU8PDgEBAQEBAQEBAQZCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwJCBgEBAQEBAQEBARISNgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUxDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUQDw4BAQEBAQEBAQEGPAICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAi0GAQEBAQEBAQEBGgw2BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/IRoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ1RBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFUw8OAQEBAQEBAQEBBkMCAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwQCSwYBAQEBAQEBAQESDDYFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFCw8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhUiBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBREKDgEBAQEBAQEBFAYZAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDBAcvDw4BAQEBAQEBAQEPEQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFADoTAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBChsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQAbDwEBAQEBAQEBARdeCAICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDCAcUFAEBAQEBAQEBDgwRBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSIMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESWSYFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgXwwBAQEBAQEBAQEXUAkCAgIIAwMDAwMDAwQIAgICAgICAgICAAAAAwMDAwMDAwMDBwICAgICAgICCQMDBy0GAQEBAQEBAQEBDz0FBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSA9DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPWgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFACQdAQEBAQEBAQEBBkUCAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAgICAgICAgICCAQEBAQEBAQECAICNAYOAQEBAQEBAQEKJSAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgDw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR1OBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT9bEwEBAQEBAQEBDgYoCQQJAgICAgICAgIHBAQEBAQEBAQEAAAAAgICAgICAgICCAMDAwMDAwMDCAIoLhoBAQEBAQEBARMkJgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFMQoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwiBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUjHhIBAQEBAQEBARReBAgJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAJIHgEBAQEBAQEBDh5PBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFP1kSAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBCjkFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFUw0BAQEBAQEBAQEGRAIJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMEAkQGAQEBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABUEeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaGD8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRsPAQEBAQEBAQEBBjQCAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwQHLwYUAQEBAQEBAQEPPQAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA6EwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPEQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT8lHQEBAQEBAQEBFF0EAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMIXR4BAQEBAQEBARMhIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVPDw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR8kIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/DBoBAQEBAQEBAQZCAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDBAJDBgEBAQEBAQEBDhQiBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFPQwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQxTBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFUw8OAQEBAQEBAQ4GSQICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDCAgSFAEBAQEBAQEBDDUFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIyEOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEzIgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTkKAQEBAQEBAQEeKgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQEBAQEAlIGAQEBAQEBAQEdVQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVMPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEOHiMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT8kHwEBAQEBAQEBBj4CAgICAgIHBAQEBAQEBAQEAAAACAgICAgICAgIBwkJCQkCKAYBAQEBAQEBAQ4YNgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUSHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQENPQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAULDA4BAQEBAQEBGiwnCAgICAgICQkJCQkJCQkJAAAAAwMDAwMDAwMEBwICAgICKx4BAQEBAQEBAQ8QBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR9ZPwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPQwBAQEBAQEBAQYtAgQDAwQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgJNBgEBAQEBAQEBEyU/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAMh0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8RBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFP1sSAQEBAQEBAQEGBAcDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAghHGgEBAQEBAQEBHjYFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIxcBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEiEmBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVwPAQEBAQEBAQEXQgkEAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAkUGAQEBAQEBAQEPUQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBRAKAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD08FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVbEwEBAQEBAQEBBhkJBAMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAl4UAQEBAQEBARoSIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMMGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKJQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2DA4BAQEBAQEBHioCAwQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICPAYBAQEBAQEBAQ8QBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVTHgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAVAoBAQEBAQEBAQYZCQQIAgICAgICAgICAAAAAwMDAwMDAwMEBwIIBhoBAQEBAQEBH1kgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUmThMBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8RBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFJhgOAQEBAQEBARcpCQMIAgICAgICAgICAAAAAwMDAwMDAwMEBwJDBgEBAQEBAQEBDFMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFQR4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGh4ABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTsMAQEBAQEBAQEGLwcIAgICAgICAgICAAAACAgICAgICAgIBwg4FAEBAQEBAQETWQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVUNAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDFMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSZbEgEBAQEBAQEXMwIIBwcHBwcHBwcHAAAAAgICAgICAgICAjQGAQEBAQEBAQEPCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSISDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETJSAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVUDAEBAQEBAQEBBhkCBAQEBAQEBAQEAAAAAgICAgICAgICAgYfAQEBAQEBARMlIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU1CgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4eIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/WxIBAQEBAQEBBkQCBAMDAwMDAwMDAAAAAgICAgICAgICGQYBAQEBAQEBAQ9TBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUmDBIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ0bBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFEA8BAQEBAQEBDgYDCAMDAwMDAwMDAAAAAgICAgICAgICVgYBAQEBAQEBGg0FBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFEQwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEh8jBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPx4OAQEBAQEBAQZKAgQDAwMDAwMDAAAAAgICAgICAgIvBgEBAQEBAQEBChEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBVUTAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDzEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVQKAQEBAQEBARReBAMDAwMDAwMDAAAAAgICAgICAgJFBgEBAQEBAQESTj8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQseAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETVT8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSIMDgEBAQEBAQEGRQkEAwMDAwMDAAAAAgICAgICAghMGgEBAQEBAQEPUQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA5CgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEeCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU5DQEBAQEBAQEeUAIDAwMDAwMDAAAAAgICAgICAkkGDgEBAQEBAQEMIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUjDw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARNOBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2DwEBAQEBAQEOHy8IAwMDAwMDAAAAAgICAgICAjAGAQEBAQEBARU9AAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFGwoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQw2BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgIRIBAQEBAQEBBkUJBAQEBAQEAAAABAQEBAQDCC4SAQEBAQEBAQxBBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPywfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDxAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFCwoBAQEBAQEBFCoJAgICAgICAAAAAwMDAwMHNAYBAQEBAQEBEls/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQsPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEODz8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIxQaAQEBAQEBAQYEAgICAgICAAAAAwMDAwQCLRcBAQEBAQEBDBsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQBZEwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMEAUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTcMAQEBAQEBAQZEAgICAgICAAAAAwMDAwMEKxQBAQEBAQEBDCMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUiCg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARM6IAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTYMAQEBAQEBARQpAgICAgICAAAAAwMDBAk0BgEBAQEBAQESTj8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFNwoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQw2BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMMAQEBAQEBAQEUKAICAgICAAAAAwMDBAIzHgEBAQEBAQEPEQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPx8OAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHTcgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUxDQEBAQEBAQEGRQICAgICAAAAAwMDAwIGGgEBAQEBAQEeIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDz8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVaHgEBAQEBAQEfWwICAgICAAAAAwMDCAQGAQEBAQEBARoKBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUKGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKPQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFHgEBAQEBAQEBBgICAgICAAAABAQECUoGAQEBAQEBAQ0QBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVTDQEBAQEBAQEBAQEBARoaGhoaGg4BAQEBAQEBAQEBAR8sIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFLB8BAQEBAQEBBhkCAgICAAAABwcJBykUAQEBAQEBAQwiBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgJQoBAQEBAQESDQ8PDxghISEhIQwPHg8dGgEBAQEBAQ8RBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFCwoBAQEBAQEBBlYCCAgIAAAAAgICJy4aAQEBAQEBEiwjBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIwwOAQESDA1AMRE/IAUFBQUFBQUgNhE9QAwMEgEBE1UABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIhgOAQEBAQEBFDgEAwMDAAAAAgICLwYBAQEBAQEBHTkFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTENEg9bGyM/BQUFBQAAAAAAAAAFBQUAPzY3FAoaDEEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFP04SAQEBAQEBDgYnAwMDAAAAAgICRAYBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMeIVM/BQUFAAAAAAAAAAAAAAAAAAAFBQUFIxEPNwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT0NAQEBAQEBAQYvCQQDAAAAAgICRx4BAQEBAQEBDyMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2PwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUjPwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABREPAQEBAQEBAQY8AgQDAAAAAgIESBQBAQEBAQEaCiAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSYPDgEBAQEBAR5HAgQDAAAAAgIvDgEBAQEBAQETJQAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSBAGgEBAQEBARRICAgDAAAAAgI0BgEBAQEBAQENEAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA5HQEBAQEBAQEGSQgDAAAAAgJFBgEBAQEBAQEKCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAURDQEBAQEBAQEGSQgDAAAACQlCFwEBAQEBAQ4TIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVBDAEBAQEBAQEGRQIIAAAABAhGHwEBAQEBARJOBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUFBQUFBQUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUiHw4BAQEBAQEXMAICAAAACAgGDgEBAQEBARNOBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBT9BPUgKDQ0NCkgbIj8FBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAWRIBAQEBAQESVwcCAAAACAQGAQEBAQEBAR0lBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUjOU4KCh8BAQEBAR8KDFU9PwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFJR0BAQEBAQEOBggCAAAABxkGAQEBAQEBAQw7BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPzEMChIBAQEBAQEBAQEBAR8MITEgBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFJR0BAQEBAQEBBgMCAAAAB0MGAQEBAQEBAQxPBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQVPEw0OAQEBAQEBAQEBAQEBAQEBGg0sTwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFNQwBAQEBAQEBBjQCAAAACUQGAQEBAQEBAQ0jBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTseHwEBAQEBAQEBAQEBAQEBAQEBAQEdFAsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFCwwBAQEBAQEBBjQCAAAACS0GAQEBAQEBDgo/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFUR4OAQEBAQEBAQEBAQEBAQEBAQEBAQEBEh4LBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIgoBAQEBAQEBBkMCAAAAAikXAQEBAQEBGiw/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVTDQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIUTwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIw8BAQEBAQEBBkQCAAAACCoUAQEBAQEBEzIgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMeEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETJD8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFP0AOAQEBAQEBFy0CAAAABCsUAQEBAQEBFRsABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFADkVAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDTEFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPzoSAQEBAQEBFykCAAAAJywaAQEBAQEBFRsABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFEQ8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwjBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFJjITAQEBAQEBFCkCAAAAFhwOAQEBAQEBFTEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/ThIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQo5BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFABsVAQEBAQEBFCsEAAAAKAYOAQEBAQEBDRAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVPCgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4BIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFABsVAQEBAQEBFCsEAAAAKAYBAQEBAQEBChAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUlHQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKUwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRAYAQEBAQEBFCsEAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUPDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfGAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRAMAQEBAQEBGhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT8PAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBCgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRAMAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVoeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHj8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRANAQEBAQEBDhwWAAAAGQYBAQEBAQEBChEFAAAAAAAAAAUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQAAAAAAAAAABREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQAAAAAAAAAABRANAQEBAQEBDhwWAAAAGQYBAQEBAQEBDBEFAAAAAAAAAAUjCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwtTPwUAAAAAAAAABREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAVPCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLPwUAAAAAAAAABRAKAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAVTHg0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ8eTwUAAAAAAAAABREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAUbDw0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ8PTwUAAAAAAAAABRAMAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAVTDQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKTwUAAAAAAAAABVoeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAUxDQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAABRAMAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAVPCgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKUwUAAAAAAAAABUEeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHgAFAAAAAAAAAAVdHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAABRANAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAVPCgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQENMQUAAAAAAAAABT8PAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEODwUAAAAAAAAAAAVAHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAABRANAQEBAQEBDhwWAAAAKAYBAQEBAQEBChAFAAAAAAAAAAVPDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfSAUAAAAAAAAAAAUNGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdJQUAAAAAAAAABQUNGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAABRAKAQEBAQEBFEwnAAAAKAYOAQEBAQEBDRAFAAAAAAAAAAVPDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDAUAAAAAAAAAAAUQCgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKUwUAAAAAAAAABT8PAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4NIgUAAAAAAAAFABsVAQEBAQEBFCsEAAAAFhwOAQEBAQEBFTEFAAAAAAAAAAUiDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHgAFAAAAAAAAAAUiIQ4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARpbIwUAAAAAAAAABUEeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4SIgUAAAAAAAAFABsVAQEBAQEBFCsEAAAAJywaAQEBAQEBFRsABQAAAAAAAAUiDg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDxEFAAAAAAAAAAUFOQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8QBQAAAAAAAAAABTEMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARJOPwUAAAAAAAAFIDITAQEBAQEBFykCAAAABCsUAQEBAQEBFRsABQAAAAAAAAUjEg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBFRsABQAAAAAAAAAFIgwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgo/BQAAAAAAAAAFPzoSAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARMyIAUAAAAAAAAFPzoSAQEBAQEBFykCAAAACCoUAQEBAQEBHT0gBQAAAAAAAAUjLBIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDiE/BQAAAAAAAAAFBTUMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDBAFAAAAAAAAAAAFIwwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR05BQAAAAAAAAAFPxMOAQEBAQEBF0ICAAAAAikXAQEBAQEBEkw/BQAAAAAAAAUgMhMBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwLBQAAAAAAAAAABSA6EwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKJQUFAAAAAAAAAAAFVAwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARU9BQAAAAAAAAAFPw0OAQEBAQEBBkQJAAAAAkIGAQEBAQEBDh0/BQAAAAAAAAUFOQoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARNZBQUAAAAAAAAAAAU2IRIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR1ZIwUAAAAAAAAAAAUmWxIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ87BQAAAAAAAAAFIwoBAQEBAQEBBlgJAAAAAkQGAQEBAQEBAQojBQAAAAAAAAAFUQ8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQENWgUAAAAAAAAAAAAFIg8fAQEBAQEBAQEBAQEBAQEBAQEBAQEBDFkjBQAAAAAAAAAABQU7DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8RBQAAAAAAAAAFCwwBAQEBAQEBBjQHAAAAAkMGAQEBAQEBAQxPBQAAAAAAAAAFOw8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETOiAFAAAAAAAAAAAABTYsDAEBAQEBAQEBAQEBAQEBAQEBAQEMVSMFAAAAAAAAAAAABSYYDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8jBQAAAAAAAAAFNQwBAQEBAQEBBjQHAAAAAhkGAQEBAQEBAQw7BQAAAAAAAAAFTw8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDxEFAAAAAAAAAAAAAAU/Mg8TAQEBAQEBAQEBAQEBAQEBDAw1IAUAAAAAAAAAAAAABTUMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwgBQAAAAAAAAAFJR0BAQEBAQEBBgQIAAAAAgQGAQEBAQEBAQxhBQAAAAAAAAAFIAwOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGg4iBQAAAAAAAAAAAAAFBQslDx8BAQEBAQEBAQEBARMMOSIFBQAAAAAAAAAAAAAFQQwOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGiEgBQAAAAAAAAAFJR0BAQEBAQEOBgMDAAAAAgMGDgEBAQEBARNOBQAAAAAAAAAFIBgaAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ1VBQUAAAAAAAAAAAAABQUgCywPHg0YFR0VGAoPD042BQUFAAAAAAAAAAAAAAU/IRIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBE2AABQAAAAAAAAUAWRIBAQEBAQESVwgDAAAAAgcuGgEBAQEBARJOBQUAAAAAAAAFACUTAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMNQUFAAAAAAAAAAAAAAAFBQUFWhAxPTkbMRBPBQUFAAAAAAAAAAAAAAAABQA5DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDREFAAAAAAAAAAUiHw4BAQEBAQEXMAkEAAAACAczHgEBAQEBARIkIwUAAAAAAAAFBRANAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDTsFBQAAAAAAAAAAAAAAAAAABQUFACAABQUFAAAAAAAAAAAAAAAAAAAFBRsMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDAsFAAAAAAAAAAVBDAEBAQEBAQEGRQIJAAAABAJFBgEBAQEBAQEKQQUAAAAAAAAABUEMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDh5TBQUAAAAAAAAAAAAAAAAAAAAABQUFAAAAAAAAAAAAAAAAAAAAAAUFOwoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESJCMFAAAAAAAAAAVTDAEBAQEBAQEGSQICAAAAAwhJBgEBAQEBAQENUwUAAAAAAAAABSMkEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIMUwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUbCg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESWQAFAAAAAAAABQA5HQEBAQEBAQEGSQICAAAAAwgoFA4BAQEBAQETJQAFAAAAAAAABQVOEwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaHzEFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTkMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETTgUAAAAAAAAABSBAGgEBAQEBARRICAICAAAAAwgESBQBAQEBAQEODSAFAAAAAAAAAAVgDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwyIwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUiLAwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMUwUAAAAAAAAABSYPDgEBAQEBAR44AgICAAAAAwMCRx4BAQEBAQEBDCMFAAAAAAAAAAULDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMWVMABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIBEBHQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEYIwUAAAAAAAAABUEPAQEBAQEBAQZWAgICAAAAAwQCRAYBAQEBAQEBDxEFAAAAAAAAAAU/IQ4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEyE/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIhcSAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR9OPwUAAAAAAAAABT0NAQEBAQEBAQY0AgICAAAAAwMHGQYBAQEBAQEBHT0FAAAAAAAAAAUAGxUBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDVMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBRsNAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ0xBQUAAAAAAAAFIDITAQEBAQEBDgYnAgICAAAAAwMDJy4aAQEBAQEBEjo/BQAAAAAAAAAFEAwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfGiAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMPDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR4RBQAAAAAAAAAFIxQOAQEBAQEBFDgEAgICAAAACAgIBykUAQEBAQEBAQoiBQAAAAAAAAAFIx4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAURDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwABQAAAAAAAAAFCw0BAQEBAQEBBlYCBwcHAAAAAgICAlYGAQEBAQEBAQ0QBQAAAAAAAAAFBRUSAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR05IAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/LBIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBE10FAAAAAAAAAAAFRh0BAQEBAQEBBk0HBAQEAAAAAgICAkkGAQEBAQEBAR8hBQAAAAAAAAAABRANAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARcjBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIh4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDUEFAAAAAAAAAAAAHgEBAQEBAQEBBgIIAwMDAAAAAgICAgIGGgEBAQEBAQEePwUAAAAAAAAABSIMDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBChAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABU4KAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESOiMFAAAAAAAAAAUiDwEBAQEBAQEaLAIDAwMDAAAAAgICAgJXHwEBAQEBAQEeEQUAAAAAAAAABQAyDQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaFCMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTYYDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKPQUAAAAAAAAAAAUxDQEBAQEBAQEGRQIEAwMDAAAAAgICAgI0BgEBAQEBAQETOSAFAAAAAAAAAAURDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUQDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEeQQUAAAAAAAAABT8hDgEBAQEBAQEULwMDAwMDAAAAAgICAgInXhQBAQEBAQEODSMFAAAAAAAAAAU/DA4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARM6AAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/HRIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR9MAAUAAAAAAAAABTYPAQEBAQEBARQpAgMDAwMDAAAAAgICAgICKRcBAQEBAQEBDBEFAAAAAAAAAAUFUQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDg9PBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFUwwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwRBQAAAAAAAAAABTcMAQEBAQEBAQZEAgQDAwMDAAAAAgICAgICSwYBAQEBAQEBH04FBQAAAAAAAAAFNgwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDzkFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBU4dAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGiQ/BQAAAAAAAAAFIxQaAQEBAQEBAQYEBwMDAwMDAAAAAgICAgICCAYOAQEBAQEBDg9PBQAAAAAAAAAFAFsfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEOISIFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABUEPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBCmAFAAAAAAAAAAAFCw0BAQEBAQEBFCoHAwQEBAQEAAAABAQEBAQECTAXAQEBAQEBAQ09BQUAAAAAAAAABUEPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPEQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA9DQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDCIFAAAAAAAAAAUAVRMBAQEBAQEBBjACAgICAgICAAAAAwMDAwMDB0MGAQEBAQEBARoYIAUAAAAAAAAABT8sEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoPIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdOSAFAAAAAAAAAAVPDwEBAQEBAQEBHy8CAgICAgICAAAAAwMDAwMDCAdIFAEBAQEBAQEPEQUAAAAAAAAAAAURHgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQxTBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFEA8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEeQQUAAAAAAAAAAAVRCgEBAQEBAQEeRgICAgICAgICAAAAAwMDAwMDBAJSFwEBAQEBAQETVSYFAAAAAAAAAAAFGA4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHTkABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPywfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARNOBQUAAAAAAAAABSMSEgEBAQEBAQEGRQICAgICAgICAAAAAwMDAwMDAwgZBgEBAQEBAQEBDAsFAAAAAAAAAAAFEQ8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHjYFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQw2BQAAAAAAAAAABREPAQEBAQEBARReBAICAgICAgICAAAAAwMDAwMDAwMJQhcBAQEBAQEBH0AFBQAAAAAAAAAFP04SAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKVAUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQBVEwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHSUgBQAAAAAAAAAFAA0aAQEBAQEBAQZWAgICAgICAgICAAAAAwMDAwMDAwQJTQYBAQEBAQEBAQ82BQAAAAAAAAAABVwPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoTIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2Cg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAAFEQ8BAQEBAQEBAQYZAgICAgICAgICAAAAAwMDAwMDAwMDAgYSAQEBAQEBAR05AAUAAAAAAAAABQBZEwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR4QBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFJQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEVMgUFAAAAAAAAAAUgVR8BAQEBAQEBHjgCAgICAgICAgICAAAABAQEBAQEBAQEAkMGAQEBAQEBAQ4NIgUAAAAAAAAAAAU2DA4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHzo/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIx8OAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoVIgUAAAAAAAAAAAUQDAEBAQEBAQEBBhkCAgICAgICAgICAAAABwcHBwcHBwcHBy8BFAEBAQEBAQEdJQUAAAAAAAAAAAUFOhUBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDwsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTEMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwbBQAAAAAAAAAABQVOEwEBAQEBAQEeOAcHCAgICAgICAgIAAAAAgICAgICAgICCAI8BgEBAQEBAQEBDVoFAAAAAAAAAAAFIxQOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETWwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUhEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBElkmBQAAAAAAAAAFBREMAQEBAQEBAQEGGQIHBAMDAwMDAwMDAAAAAgICAgICAgICCAgEBg4BAQEBAQEBEzogBQAAAAAAAAAABRAPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4VNgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAULDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD1oFAAAAAAAAAAAFIAwSAQEBAQEBAR9OBwIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQCVgYBAQEBAQEBAQ9TBQAAAAAAAAAABQUhHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ81BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAMg0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdTgUFAAAAAAAAAAUFEAwBAQEBAQEBAQZDAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQIBBoUAQEBAQEBARJZPwUAAAAAAAAAAAUiDw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGgwmBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFNhgOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIkIwUAAAAAAAAAAAUiHRoBAQEBAQEBGjoCAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQDAjAGAQEBAQEBAQEPUQUAAAAAAAAAAAUFNQwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDBAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBT0NAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8RBQAAAAAAAAAAAAUyCgEBAQEBAQEBBkMCAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDCAQkFAEBAQEBAQEBHiMFAAAAAAAAAAAFAE4dAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfWQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMTEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDFQFBQAAAAAAAAAABUEeAQEBAQEBAQEeRwgCAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDBAc8BgEBAQEBAQEBHTkgBQAAAAAAAAAABTYNDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUQDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfJCMFAAAAAAAAAAAFP04TAQEBAQEBARoGGQICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQDAwMCBhoBAQEBAQEBAQ9aBQAAAAAAAAAAAAVTDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQo5AAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgVR8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4MNgUAAAAAAAAAAAAFEAwBAQEBAQEBAR44AgICAgIHBAMDAwMDAwMDAAAACQkJCQkJCQkJCAgICAgCNAYBAQEBAQEBARofIwUAAAAAAAAAAAUgOQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDg8jBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFQR4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ9TBQAAAAAAAAAAAAU2DQ4BAQEBAQEBGgEoAgkJCQkHCAgICAgICAgIAAAABAQEBAQEBAQEBwICAgICAjMXAQEBAQEBAQEMVAUFAAAAAAAAAAAFJgofAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDRsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBSEfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD1UFBQAAAAAAAAAABQVVHQEBAQEBAQEBBkQCBAQEBAQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAigGDgEBAQEBAQEBDxEFAAAAAAAAAAAABT8UHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESFCMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMPDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdTj8FAAAAAAAAAAAABTUMAQEBAQEBAQEUDgkIAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgJWBgEBAQEBAQEBHywjBQAAAAAAAAAABQVTHhoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPUwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUbDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIPIwUAAAAAAAAAAAAFIg8OAQEBAQEBAQ4GQwIEAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgIILBQBAQEBAQEBAQ1OBQUAAAAAAAAAAAUFEQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARNZAAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/DA4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDh5BBQAAAAAAAAAAAAUgIR8BAQEBAQEBAQZCAgMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICGQYBAQEBAQEBAQEPEAUFAAAAAAAAAAAFBRsMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgw2BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFEQ8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDTsFBQAAAAAAAAAABQA9DwEBAQEBAQEBFF4EBwQDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICAi0GAQEBAQEBAQEOGDYFAAAAAAAAAAAABQA5DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHToFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFAFsfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPNQUFAAAAAAAAAAAABREMAQEBAQEBAQEBBksCBAMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICAgJIFAEBAQEBAQEBHx4/BQAAAAAAAAAAAAUAVQwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEOIT8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABUEMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQw9BQUAAAAAAAAAAAAFTw8OAQEBAQEBAQEGMwIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQkZBg4BAQEBAQEBAR1OIAUAAAAAAAAAAAAFPzIMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPOwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA5HQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEODxsFBQAAAAAAAAAAAAUmJBMBAQEBAQEBARReLwIIAwMDAwMDAwMIAgICAgICAgICAAAABAQEBAQEBAQEBwICAgICAgICCQQCVgYBAQEBAQEBAQEMVAAFAAAAAAAAAAAABSBVDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoKIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUjHgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoPPQUFAAAAAAAAAAAABSAlCgEBAQEBAQEBDgY0AgIIBAQEBAQEBAQIAgICAgICAgICAAAAAgICAgICAgICCAMDAwMDAwMDCAICB1AeAQEBAQEBAQEBDBAFBQAAAAAAAAAAAAUgMgwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQxTBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFXwwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwbBQUAAAAAAAAAAAAFBUYNAQEBAQEBAQEBBi0HAwMJAgICAgICAgIHAwMDAwMDAwMDAAAAAgICAgICAgICCAQDAwMDAwMDCAICAgQXGgEBAQEBAQEBDg9BBQUAAAAAAAAAAAAFP1UMAQEBAQEBAQEBAQEBAQEBAQEBAQEBCiUgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIxQSAQEBAQEBAQEBAQEBAQEBAQEBAQEBDz0ABQAAAAAAAAAAAAUFER4BAQEBAQEBAQEeWQgIAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgI0BgEBAQEBAQEBAQEMQQUAAAAAAAAAAAAABQA5DAEBAQEBAQEBAQEBAQEBAQEBAQEODEEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVMPAQEBAQEBAQEBAQEBAQEBAQEBAQ4NNQUFAAAAAAAAAAAABQURDwEBAQEBAQEBARQPBAcEAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICPAYBAQEBAQEBAQESFSIFAAAAAAAAAAAAAAUAGwoOAQEBAQEBAQEBAQEBAQEBAQEVOQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSA6HwEBAQEBAQEBAQEBAQEBAQEBEh47BQUAAAAAAAAAAAAFBTYMEgEBAQEBAQEBGgZLBwQDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAisGAQEBAQEBAQEBEiEiBQAAAAAAAAAAAAAFBTsMEgEBAQEBAQEBAQEBAQEBAQ4PIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVBDQEBAQEBAQEBAQEBAQEBAQETD1MFBQAAAAAAAAAAAAAFNgwaAQEBAQEBAQEBBkQCAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgQ4HwEBAQEBAQEBARNOIgUAAAAAAAAAAAAABQVTFB8BAQEBAQEBAQEBAQEBAQ87BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFOQoBAQEBAQEBAQEBAQEBAR0sIgUFAAAAAAAAAAAAAAU2EhIBAQEBAQEBAQEGQgIDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgIWBhQBAQEBAQEBAQETISYFAAAAAAAAAAAAAAUFIgoNAQEBAQEBAQEBAQEBElkmBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIwwBAQEBAQEBAQEBAQEBCiE/BQUAAAAAAAAAAAAABSMPEgEBAQEBAQEBARcrBAgDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICKAYUAQEBAQEBAQEBEiEiBQAAAAAAAAAAAAAABSY5Dw4BAQEBAQEBAQEBHgsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTsKAQEBAQEBAQEBARIPMQAFAAAAAAAAAAAAAAAFNg8OAQEBAQEBAQEBHlknCQQDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAjQGAQEBAQEBAQEBARNOIgUAAAAAAAAAAAAAAAUgEA0TAQEBAQEBAQEfHwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSAeEgEBAQEBAQEBCh9aBQUAAAAAAAAAAAAABQU2EhIBAQEBAQEBAQEfBigCBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAABwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwI+BgEBAQEBAQEBAQETISIFBQAAAAAAAAAAAAAFBUEUDA4BAQEBAQENNgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVTDAEBAQEBAQ4MVSMFBQAAAAAAAAAAAAAFBTYMEgEBAQEBAQEBAQEGKAIIBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEBAICVh4BAQEBAQEBAQEBEhVBBQUAAAAAAAAAAAAAAAUAVB4SAQEBAQwbBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUmVRMBAQEBHQ81BQUAAAAAAAAAAAAAAAUFEQwaAQEBAQEBAQEBAQY0AgMCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAlIGAQEBAQEBAQEBARIMQQUFAAAAAAAAAAAAAAAFAE8hDw4BDh4/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIh4OAQ4PTiIFBQAAAAAAAAAAAAAABQURDxIBAQEBAQEBAQEBBlYHBAMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgJSBg4BAQEBAQEBAQEBDxAABQAAAAAAAAAAAAAABQUgOxgTDxAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBVQMHRQRBQUFAAAAAAAAAAAAAAAFID0eAQEBAQEBAQEBARQGPgIDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgIJUgYBAQEBAQEBAQEBDgxUIAUAAAAAAAAAAAAAAAAFBSIsJD8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMPOiMFBQAAAAAAAAAAAAAAAAUmJQ0OAQEBAQEBAQEBAQZWBwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICCVIGAQEBAQEBAQEBAQEMTj8FBQAAAAAAAAAAAAAABQUAIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUiBQUFAAAAAAAAAAAAAAAFBU8kCgEBAQEBAQEBAQEOBj4HAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgJSHgEBAQEBAQEBAQEBHR42BQUAAAAAAAAAAAAAAAAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQAAAAAAAAAAAAAAAAUAER4TAQEBAQEBAQEBAQEGVgIDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICVgYBAQEBAQEBAQEBAR8YEAUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBSA9DA4BAQEBAQEBAQEBDgZWBwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAj4GFAEBAQEBAQEBAQEOD04jBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFIiEPAQEBAQEBAQEBAQEaBhkCBAMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwQIAgICAgICAgICAAAAAwMDAwMDAwMDBwICAgICAgICCQMDAwMDAwMDAwICAgICAgICAgI0BhQBAQEBAQEBAQEBARUsEQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQU1FB8BAQEBAQEBAQEBARcuGQIDAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAgMEBAQEBAQEBAcCKAYfAQEBAQEBAQEBAQEfD1QjBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFNlUPGgEBAQEBAQEBAQEBBisnAggCAgICAgICAgIEBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAhY4BgEBAQEBAQEBAQEBAQwfWiAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT8QDQoBAQEBAQEBAQEBAQEGQgQJBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgIEKwYBAQEBAQEBAQEBAQESDDkjBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFNiUPDgEBAQEBAQEBAQEBDgYtAggEBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAjwGGgEBAQEBAQEBAQEBAR0eUT8FBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBSNRHhMBAQEBAQEBAQEBAQEUD0sCAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgI0Fx4BAQEBAQEBAQEBAQEBD1lTIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUgEBIKAQEBAQEBAQEBAQEBAQYULwIDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICBFAGDgEBAQEBAQEBAQEBARIPOloFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBVo6DxoBAQEBAQEBAQEBAQEaBi0HBwQDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgdWBhQBAQEBAQEBAQEBAQEBEw0lIgAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBSALWQ0TAQEBAQEBAQEBAQEBARQuNAIIBAMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgICGUgGAQEBAQEBAQEBAQEBAQEdDTk2BQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFUyUKHwEBAQEBAQEBAQEBAQEBBkgoAgMEAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAgMEBAQEBAQEBAcCAgICAgICAgIECQItBh4BAQEBAQEBAQEBAQEBDh0PQAsmBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFPxENDxMBAQEBAQEBAQEBAQEBAR4GWAICAgMEBAQEBAQEBAcCAgICAgICAgIEBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAACAgICAgICAgIBwkJCQkJCQkJBwgICAgICAgICAkJCQkJCQkJCQcICAgICAgICAgJCQICKEYGDgEBAQEBAQEBAQEBAQEBHwxVESYFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFACNRTgoaAQEBAQEBAQEBAQEBAQEUBl0vBwgICAkJCQkJCQkJCQcICAgICAgICAgJCQkJCQkJCQkICAgICAgICAgICQkJCQkJCQkJAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAghKBhcBAQEBAQEBAQEBAQEBAQETDyE9NgUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBT9BPRgPEgEBAQEBAQEBAQEBAQEBDgYSQwgHBAMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIBAMDAwMDAwQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICKDMGHwEBAQEBAQEBAQEBAQEBAQ4ND04RIyAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBT8jG1sMDQ4BAQEBAQEBAQEBAQEBAQEeBlIIAgMEAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgI0BgYUAQEBAQEBAQEBAQEBAQEBDh8PGDkRPwUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUiEU4MDBIBAQEBAQEBAQEBAQEBAQEBHwYrGQcIBAMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgICAjwkBhQBAQEBAQEBAQEBAQEBAQEBDhMPHkAQIiMFBQUFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUFBSMiEAoeChIBAQEBAQEBAQEBAQEBAQEBARQGR00CAgQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwcEMBoGDgEBAQEBAQEBAQEBAQEBAQEBARINCjo9ESMgAAUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUAICMROSwMDRoBAQEBAQEBAQEBAQEBAQEBAQEaBl5FCAICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAgRWBgYUAQEBAQEBAQEBAQEBAQEBAQEBARIdDwwNJVNBIwUFBQUFBQUFBQUFBQAAAAAABQUFBQAAAAAFBQUFBQUFBQUFBQUjCxAlDQwPHRIBAQEBAQEBAQEBAQEBAQEBAQEBFAYGSgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwgCBDwBBhIBAQEBAQEBAQEBAQEBAQEBAQEBAQEOEw0YJE5OJTtPIz8/IAAABQUFBQUFBQUFBQUFBQUFAAAgPz8jTzslTk4kCg0TDgEBAQEBAQEBAQEBAQEBAQEBAQEBAR8GOEMIAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDBAMCAgICAgICAgIIBAMDBAQECAIvQwYGFwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEhIfHQwMDR1MPRsbMRAQEREREREREREREBAxGxsyTBUKDAwKExISAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEXBgY0CAICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDBAMCAgICAgICAgIIBAMDAwMDAwQIAgICAgICAgICAAAACAgICAgICAgIBwcHBwcHBwcHBwgICAgICAgICAcHBwcHBwcHBwcICAgICAgICAgHBwcHBwcHBwcICAgICAgICAgHAgJNQgYGFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4SHRUVFRgYDAwPDw8PDwwMGBgVFRUTEg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4aBgZCGQICBwcHBwcHBwcHBwgICAgICAgICAcHBwcHBwcHBwcICAgICAgICAgHBwcHBwcHBwcICAgICAgICAgIBwcHBwcHBwcHAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAgMEBAQEBAQEBAcCAgICAgICAgIEBAQEBAQEBAQJAgICAgICAgIHBAMJCS9ESAYXDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaBgZMRS8JAgICCAQEBAQEBAQECAICAgICAgICAgMEBAQEBAQEBAcCAgICAgICAgIEBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMEAwgCB0MwBgYXFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEUBgYuMEkIAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMECAcJCEspXgYeGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHgYrLTQIAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICJzQqBgYGFBoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIXBgYGMzQEAgcDAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAklWKS4GBh4UDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBFB4GBi4pSgQCAgkDBAMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICBycZREdIFAYGHhoOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDh8XBgYOSEdELycHCQgIBAQDAwMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMHAgIEKElFMy4GBgYGBhcXFBQaEg4BAQEBAQEBAQEBAQ4SGhQUFwYGBgYGBkZCRTQvBAICAgIJBAMDAwMDAwMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMDBAMICAgCBwcDBBlDREIpKissHAYGBgYGBgYGBgYGBgYcLCsqKS1EQxkECAgJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMDAwMDAwMECAICAgICAgICCAQnFigoGRkZGRkZGRkZKCgWJwQIAgkJBwcICAQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"""
icondata= base64.b64decode(icon)
## The temp file is icon.ico
tempFile= "icon.ico"
iconfile= open(tempFile,"wb")
## Extract the icon
iconfile.write(icondata)
iconfile.close()
root.wm_iconbitmap(tempFile)
#top.wm_iconbitmap(tempFile)

## Delete the tempfile
os.remove(tempFile)
widthh = root.winfo_screenwidth()
heighttt= root.winfo_screenheight()
t =widthh/2
b =heighttt/2

#top.title("CONFIGURATION")
root.title("CN_REPORTER")
root.geometry("370x250+1000+100")
root.call('wm', 'attributes', '.', '-topmost', True)
#root.overrideredirect(1)
root.resizable(False,False)
root.protocol('WM_DELETE_WINDOW', doSomething)
#root.attributes('-disabled', True)
root.mainloop()



